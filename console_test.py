"""
Synthetic verification for the hours-based discipline block + Budgets (Input) loop.
No DB. Exercises: seed Budgets form → simulate PM fill → read back → assemble the
dashboard → assert discipline % = ETO actual hrs ÷ PM budget hrs. Renders both sheets.
"""
import datetime as dt
import pandas as pd
from openpyxl import load_workbook

import console_dashboard as cd
import console_budgets as cb
from console_feed import DISCIPLINES

# ── 1. Synthetic Layer-1 summary (now carries ETO 3-bucket estimate hours) ──────
summary = pd.DataFrame([
    dict(ProjectID=230219, Project="230219 - 5500T Forging Press",
         EstAdminHours=4000, EstEngHours=12000, EstMfgHours=28000,
         LabEstHrs=44000, LabActHrs=52458, LabPctHrs=1.249,
         LabEstCost=3696000, LabActCost=4616000, LabPctCost=1.267,
         MatEst=5100000, MatAct=4697100, MatPct=0.922,
         TotalBudget=8796000, ActTotalCost=9313100, TotalPct=1.059,
         SalesPrice=10200000, BudgetMargin=0.138, ActualMargin=0.087),
    dict(ProjectID=240033, Project="240033 - 3000T Trim Press",
         EstAdminHours=2500, EstEngHours=9000, EstMfgHours=16500,
         LabEstHrs=28000, LabActHrs=21560, LabPctHrs=0.770,
         LabEstCost=2464000, LabActCost=1897000, LabPctCost=0.770,
         MatEst=3200000, MatAct=2240000, MatPct=0.700,
         TotalBudget=5664000, ActTotalCost=4137000, TotalPct=0.730,
         SalesPrice=6800000, BudgetMargin=0.167, ActualMargin=0.208),
])

# ── 2. Synthetic ACTUAL hours by discipline (ETO numerator) ─────────────────────
def _disc(pid, hrs):
    return [dict(ProjectID=pid, Discipline=d, ActHours=h)
            for d, h in zip(DISCIPLINES, hrs)]
disc = pd.DataFrame(
    _disc(230219, [5000, 8000, 3000, 4000, 30000, 200]) +
    _disc(240033, [2200, 6500, 2100, 3300, 9000, 60])
)

# ── 3. Seed the Budgets (Input) form from ETO, then simulate the PM filling it ──
cb.write_budgets_input("budgets_input.xlsx", summary)
wb = load_workbook("budgets_input.xlsx")
ws = wb[cb.BUDGETS_SHEET]
hdr = cb._find_header_row(ws)
col = {ws.cell(hdr, c).value: c for c in range(1, ws.max_column + 1)}
def setcell(pid, header, val):
    for r in range(hdr + 1, ws.max_row + 1):
        if str(ws.cell(r, col["Project ID"]).value).split(".")[0] == str(pid):
            ws.cell(r, col[header]).value = val
# PM splits Engineering (12000) into Mech/Hyd/Elec + enters dates for 230219
setcell(230219, "Budget Mechanical Hrs", 6000)
setcell(230219, "Budget Hydraulic Hrs", 2500)
setcell(230219, "Budget Electrical Hrs", 4500)   # Mech+Hyd+Elec = 13000 (Eng bucket = 12000)
setcell(230219, "PO Ship Date", dt.datetime(2026, 5, 15))
setcell(230219, "Customer Agreed Ship Date", dt.datetime(2026, 7, 31))
setcell(230219, "Planned Ship Date", dt.datetime(2026, 9, 18))
setcell(230219, "% Done", 0.82)
# 240033: leave discipline split blank (tests the blank-budget path) but add dates
setcell(240033, "PO Ship Date", dt.datetime(2026, 11, 30))
setcell(240033, "Customer Agreed Ship Date", dt.datetime(2026, 11, 30))
setcell(240033, "Planned Ship Date", dt.datetime(2026, 11, 24))
setcell(240033, "% Done", 0.55)
wb.save("budgets_input.xlsx")

# ── 4. Read it back into an overlay + a small CSV overlay for procurement ───────
overlay_budgets = cb.read_budgets_input("budgets_input.xlsx")
overlay_csv = pd.DataFrame([
    dict(ProjectID="230219", Rank=1, TotalLineItems=612, LLTPOrdered=48, LLTPRelLate=6),
])
overlay = cd._merge_overlays([overlay_budgets, overlay_csv])

# ── 5. 2-week labour + assemble ─────────────────────────────────────────────────
today = dt.date(2026, 7, 24)
labor = pd.DataFrame([{"Project ID": "230219", "Actual Hours": 40, "Date": today - dt.timedelta(d)}
                      for d in range(14)] +
                     [{"Project ID": "240033", "Actual Hours": 25, "Date": today - dt.timedelta(d)}
                      for d in range(14)])
tw = cd.two_week_labour_hours(labor, as_of=today)
df = cd.build_executive_dashboard(summary, disc, overlay, two_week=tw)

# ── Assertions ──────────────────────────────────────────────────────────────────
r219 = df[df["ProjectID"] == "230219"].iloc[0]
r033 = df[df["ProjectID"] == "240033"].iloc[0]

# discipline % = ETO actual hrs ÷ PM budget hrs
assert abs(r219["disc::Electrical Engineering"] - 4000/4500) < 1e-6, r219["disc::Electrical Engineering"]
assert abs(r219["disc::Mechanical Engineering"] - 8000/6000) < 1e-6, r219["disc::Mechanical Engineering"]
assert abs(r219["disc::Project Management"] - 5000/4000) < 1e-6   # PM seed = EstAdmin 4000
# 240033 has no discipline budgets entered → discipline % blank
assert r033["disc::Electrical Engineering"] is None or pd.isna(r033["disc::Electrical Engineering"])
# schedule sourced from Budgets input
assert r219["SlippageDays"] == 49, r219["SlippageDays"]
assert abs(r219["RunoutLabour"] - (1.249/0.82)) < 1e-6
assert r219["PctDone"] == 0.82
# procurement from CSV overlay merged in
assert r219["TotalLineItems"] == 612
# ETO budget block unchanged
assert abs(r219["LabPctHrs"] - 1.249) < 1e-9 and abs(r219["MatPct"] - 0.922) < 1e-9
print("All hours-basis + Budgets-loop assertions passed.")

print("\nDiscipline block (230219, hours basis = actual ÷ budget):")
for d in cd._DISC_ORDER:
    v = r219[f"disc::{d}"]
    print(f"  {d:<24} {'' if v is None or pd.isna(v) else f'{v:.1%}'}")

# ── Render both sheets ──────────────────────────────────────────────────────────
cd.write_executive_dashboard(df, "sample_dashboard.xlsx", as_of=today)
print("\nWrote budgets_input.xlsx and sample_dashboard.xlsx")
