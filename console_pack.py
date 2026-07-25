"""
console_pack.py
================
Readers for the REAL management pack's two manual entry points — the way PMs
already maintain them. No re-keying, no seeded form: the dashboard reads these.

  * Budgets tab       — the plan: PO/Customer ship dates, per-discipline budget
                        hours (the 'Helpers' roll-up AW:BA), material & labour totals.
                        Its column grouping IS the authoritative HourDescription→
                        discipline crosswalk (derive_crosswalk_from_budgets).
  * PM Entries tab    — weekly time series: Planned Ship, % Done, run-out %s,
                        material actual/budget, procurement block, 1-week deltas,
                        Rank/Re-Rank. One row per project per week; we take the
                        'Front Page Selection' week (or a given week).

Verified against the live pack 2026-07-25 (Budgets 101×56, PM Entries 501×28).
Read by header/known columns; returns overlay-shaped frames the dashboard consumes.
"""
from __future__ import annotations

import datetime as _dt
import pandas as pd
from openpyxl import load_workbook

BUDGETS_SHEET = "Budgets"
PM_SHEET = "PM Entries"

# Budgets tab (1-based columns). Data begins row 6; headers row 5, groups row 4.
_B_ID, _B_PROJ, _B_POSHIP, _B_CUSTAGREED = 2, 3, 4, 5
_B_MAT_TOTAL, _B_LAB_TOTAL = 8, 56
# 'Helpers' discipline roll-ups
_B_HELPERS = {
    "Project Management":     49,  # AW
    "Mechanical Engineering": 50,  # AX
    "Electrical Engineering": 51,  # AY
    "Hydraulic Engineering":  52,  # AZ
    "Manufacturing":          53,  # BA
}
_B_DETAIL_FIRST, _B_DETAIL_LAST = 9, 47   # I..AU  (HourDescription budget columns)
_B_GROUP_ROW, _B_HEAD_ROW, _B_DATA_ROW = 4, 5, 6

_DISC_NAMES = {"Project Management", "Mechanical Engineering", "Electrical Engineering",
               "Hydraulic Engineering", "Manufacturing", "Other"}

# PM Entries tab (1-based). Data from row 13; header row 12; week selector at C8.
_PM = dict(ID=2, CLIENT=3, YEAR=4, WEEK=5, PLANNED_SHIP=6, PCT_DONE=7, LAB_RUNOUT=8,
           MAT_RUNOUT=9, MAT_ACTUAL=10, LINE_ITEMS=11, LLTP_ORD=12, LLTP_REL_LATE=13,
           LLTP_ORD_LATE=14, LLTP_DEL_LATE=15, PARTS_REL_LATE=16, PARTS_ORD_LATE=17,
           INCLUDE=18, YEARWEEK=19, YEARWEEKKEY=20, D_PCTDONE=21, D_MAT=22,
           MAT_BUDGET=23, RANK=24, RERANK=25)
_PM_WEEK_CELL = (8, 3)   # C8 'Front Page Selection'
_PM_DATA_ROW = 13


def _pid(v):
    if v is None or isinstance(v, str) and not v.strip():
        return None
    try:
        return str(int(float(v)))
    except (TypeError, ValueError):
        return None


# ─────────────────────────────────────────────────────────────────────────────
# Budgets tab
# ─────────────────────────────────────────────────────────────────────────────
def read_budgets_tab(path):
    """Per-project plan from the Budgets tab → overlay-shaped frame:
    ProjectID, POShipDate, CustAgreedDate, MatBudgetTotal, LabBudgetTotal,
    BudgetHrs::<discipline> (the 5 Helpers roll-ups)."""
    wb = load_workbook(path, data_only=True)
    ws = wb[BUDGETS_SHEET]
    rows = []
    for r in range(_B_DATA_ROW, ws.max_row + 1):
        pid = _pid(ws.cell(r, _B_ID).value)
        if pid is None:
            continue
        row = {
            "ProjectID": pid,
            "POShipDate": ws.cell(r, _B_POSHIP).value,
            "CustAgreedDate": ws.cell(r, _B_CUSTAGREED).value,
            "MatBudgetTotal": ws.cell(r, _B_MAT_TOTAL).value,
            "LabBudgetTotal": ws.cell(r, _B_LAB_TOTAL).value,
        }
        for disc, col in _B_HELPERS.items():
            row[f"BudgetHrs::{disc}"] = ws.cell(r, col).value
        rows.append(row)
    return pd.DataFrame(rows)


def read_budgets_detail(path):
    """Fine-grain budget hours per project × HourDescription (Budgets cols I:AU).
    Long form: ProjectID, HourDescription, BudgetHours (skips blanks/zeros)."""
    wb = load_workbook(path, data_only=True)
    ws = wb[BUDGETS_SHEET]
    headers = {c: ws.cell(_B_HEAD_ROW, c).value for c in range(_B_DETAIL_FIRST, _B_DETAIL_LAST + 1)}
    rows = []
    for r in range(_B_DATA_ROW, ws.max_row + 1):
        pid = _pid(ws.cell(r, _B_ID).value)
        if pid is None:
            continue
        for c, hd in headers.items():
            if not hd:
                continue
            v = ws.cell(r, c).value
            if v in (None, "", 0):
                continue
            rows.append({"ProjectID": pid, "HourDescription": str(hd).strip(),
                         "BudgetHours": v})
    return pd.DataFrame(rows)


def derive_crosswalk_from_budgets(path):
    """Build HourDescription→discipline from the Budgets tab's own column grouping
    (row 4 group headers forward-filled across the detail columns row 5). This makes
    the Budgets tab the single source of truth, so actuals and budgets always agree."""
    wb = load_workbook(path, data_only=True)
    ws = wb[BUDGETS_SHEET]
    group = None
    xwalk = {}
    for c in range(_B_DETAIL_FIRST, _B_DETAIL_LAST + 1):
        g = ws.cell(_B_GROUP_ROW, c).value
        if g in _DISC_NAMES:
            group = g
        detail = ws.cell(_B_HEAD_ROW, c).value
        if detail and group:
            xwalk[str(detail).strip()] = group
    return xwalk


# ─────────────────────────────────────────────────────────────────────────────
# PM Entries tab
# ─────────────────────────────────────────────────────────────────────────────
def selected_week(path):
    wb = load_workbook(path, data_only=True)
    ws = wb[PM_SHEET]
    return ws.cell(*_PM_WEEK_CELL).value


def read_pm_entries(path, week=None, all_weeks=False):
    """Per-project progress/material/procurement. Default: the pack's 'Front Page
    Selection' week. all_weeks=True returns every weekly row (for history backfill)."""
    wb = load_workbook(path, data_only=True)
    ws = wb[PM_SHEET]
    if week is None:
        week = ws.cell(*_PM_WEEK_CELL).value
    rows = []
    for r in range(_PM_DATA_ROW, ws.max_row + 1):
        pid = _pid(ws.cell(r, _PM["ID"]).value)
        if pid is None:
            continue
        yw = ws.cell(r, _PM["YEARWEEK"]).value
        if not all_weeks and week is not None and str(yw) != str(week):
            continue
        inc = ws.cell(r, _PM["INCLUDE"]).value
        rows.append({
            "Year": ws.cell(r, _PM["YEAR"]).value,
            "WeekNo": ws.cell(r, _PM["WEEK"]).value,
            "YearWeek": yw,
            "YearWeekKey": ws.cell(r, _PM["YEARWEEKKEY"]).value,
            "ProjectID": pid,
            "PlannedShipDate": ws.cell(r, _PM["PLANNED_SHIP"]).value,
            "PctDone": ws.cell(r, _PM["PCT_DONE"]).value,
            "RunoutLabour": ws.cell(r, _PM["LAB_RUNOUT"]).value,
            "RunoutMaterial": ws.cell(r, _PM["MAT_RUNOUT"]).value,
            "MatActual": ws.cell(r, _PM["MAT_ACTUAL"]).value,
            "MatBudget": ws.cell(r, _PM["MAT_BUDGET"]).value,
            "TotalLineItems": ws.cell(r, _PM["LINE_ITEMS"]).value,
            "LLTPOrdered": ws.cell(r, _PM["LLTP_ORD"]).value,
            "LLTPRelLate": ws.cell(r, _PM["LLTP_REL_LATE"]).value,
            "LLTPOrdLate": ws.cell(r, _PM["LLTP_ORD_LATE"]).value,
            "LLTPDelLate": ws.cell(r, _PM["LLTP_DEL_LATE"]).value,
            "PartsRelLate": ws.cell(r, _PM["PARTS_REL_LATE"]).value,
            "PartsOrdLate": ws.cell(r, _PM["PARTS_ORD_LATE"]).value,
            "PctDoneDelta": ws.cell(r, _PM["D_PCTDONE"]).value,   # 1-week in the pack
            "MatSpend2wk": ws.cell(r, _PM["D_MAT"]).value,        # 1-week materials delta
            "Rank": ws.cell(r, _PM["RERANK"]).value or ws.cell(r, _PM["RANK"]).value,
            "Include": inc,
        })
    return pd.DataFrame(rows)


def read_pack_overlay(path, week=None):
    """Merge Budgets + PM Entries into one overlay frame (one row per ProjectID).
    PM Entries values take precedence where both carry a field (e.g. ship dates)."""
    b = read_budgets_tab(path)
    p = read_pm_entries(path, week=week)
    if b.empty and p.empty:
        return None
    df = pd.concat([b, p], ignore_index=True)
    return df.groupby("ProjectID", as_index=False).first()
