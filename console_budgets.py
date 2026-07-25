"""
console_budgets.py
===================
The Budgets tab as the PM ENTRY POINT for the manual half of the Project Console dashboard.

ETO can source everything except the per-discipline budget SPLIT and the schedule.
This module makes the Budgets tab a smart input form and closes the loop:

    ETO seeds  →  PM splits Engineering + enters dates  →  ETO reads back  →  dashboard

`write_budgets_input()` writes a `Budgets (Input)` sheet SEEDED from ETO — project
list, ETO's 3-bucket estimate hours (Admin/Eng/Mfg) and total budget hours as
guardrails, with PM's discipline budget cells pre-filled where ETO already knows
them (Admin→PM, Mfg→Manufacturing) and BLANK for the three the PM must split out of
Engineering (Mechanical / Hydraulic / Electrical). A live reconciliation column
(Σ entered − ETO total) flags a mis-key as the PM types.

`read_budgets_input()` reads that sheet back into an overlay-shaped frame the
dashboard consumes: budget hours per discipline (the denominator of the hours-based
discipline block) + the ship dates + % done.

Design notes
  * Budget hours per discipline are NOT in ETO at 6-discipline grain — only Admin/
    Eng/Mfg exist. This form is where the finer split is authored. (Phase-plan Q:
    "where do discipline budgets live long-term" — answer for now: here.)
  * Read is by HEADER NAME, not fixed cell refs, so inserting/reordering columns
    won't silently break it.
"""
from __future__ import annotations

import datetime as _dt

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BUDGETS_SHEET = "Budgets (Input)"

# Palette (aligned with console_dashboard)
C_TITLE = "1F3864"
C_SEED  = "E7E6E6"   # ETO-seeded reference (don't edit) — grey
C_ENTRY = "FFF2CC"   # PM entry — light yellow
C_CALC  = "DDEBF7"   # live formula (reconciliation) — light blue
C_HEAD  = "1F3864"
WHITE   = "FFFFFF"
GREY    = "808080"

# The 5 disciplines a PM budgets against (Other carries no budget).
BUDGET_DISCIPLINES = [
    ("Project Management",     "Budget PM Hrs"),
    ("Mechanical Engineering", "Budget Mechanical Hrs"),
    ("Hydraulic Engineering",  "Budget Hydraulic Hrs"),
    ("Electrical Engineering", "Budget Electrical Hrs"),
    ("Manufacturing",          "Budget Manufacturing Hrs"),
]

# Column layout (header -> role). Order defines the sheet columns.
#   seed  = ETO-filled reference (locked look)
#   entry = PM types here
#   calc  = Excel formula
_LAYOUT = [
    ("Project ID",                "seed"),
    ("Project",                   "seed"),
    ("ETO Est Admin Hrs",         "seed"),
    ("ETO Est Eng Hrs",           "seed"),
    ("ETO Est Mfg Hrs",           "seed"),
    ("ETO Total Budget Hrs",      "seed"),
    ("Budget PM Hrs",             "entry"),
    ("Budget Mechanical Hrs",     "entry"),
    ("Budget Hydraulic Hrs",      "entry"),
    ("Budget Electrical Hrs",     "entry"),
    ("Budget Manufacturing Hrs",  "entry"),
    ("Σ Entered Hrs",             "calc"),
    ("Δ vs ETO Total",            "calc"),
    ("PO Ship Date",              "entry"),
    ("Customer Agreed Ship Date", "entry"),
    ("Planned Ship Date",         "entry"),
    ("% Done",                    "entry"),
    ("Material Budget $",         "entry"),
]
_HEADERS = [h for h, _ in _LAYOUT]

# read-back mapping: sheet header -> overlay key the dashboard understands
_READ_MAP = {
    "Budget PM Hrs":              "BudgetHrs::Project Management",
    "Budget Mechanical Hrs":      "BudgetHrs::Mechanical Engineering",
    "Budget Hydraulic Hrs":       "BudgetHrs::Hydraulic Engineering",
    "Budget Electrical Hrs":      "BudgetHrs::Electrical Engineering",
    "Budget Manufacturing Hrs":   "BudgetHrs::Manufacturing",
    "PO Ship Date":               "POShipDate",
    "Customer Agreed Ship Date":  "CustAgreedDate",
    "Planned Ship Date":          "PlannedShipDate",
    "% Done":                     "PctDone",
    "Material Budget $":          "MatBudget",
}

TITLE_ROW, HEAD_ROW, DATA_ROW = 1, 2, 3


def _col(header):
    return _HEADERS.index(header) + 1


def write_budgets_input(output_path, seed_df, template_path=None):
    """
    Write/refresh the seeded `Budgets (Input)` sheet.

    seed_df: one row per project with columns ProjectID, Project, EstAdminHours,
             EstEngHours, EstMfgHours, and (optionally) LabEstHrs / TotalBudgetHrs.
             Comes straight from console_engine.build_project_summary().

    If `template_path` is given the sheet is added to a copy of that workbook (so an
    existing Budgets (Input) is REPLACED but any PM-entered values on it are carried
    forward first — we never silently wipe entered data). Otherwise a fresh workbook.
    """
    prior = _read_entries(template_path) if template_path else {}

    if template_path:
        wb = load_workbook(template_path)
        if BUDGETS_SHEET in wb.sheetnames:
            del wb[BUDGETS_SHEET]
        ws = wb.create_sheet(BUDGETS_SHEET)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = BUDGETS_SHEET

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.cell(TITLE_ROW, 1, "Project Console Budgets — PM Entry (ETO seeds grey · you fill yellow)"
            ).font = Font(name="Helvetica", size=11, bold=True, color=C_TITLE)

    head_fill = PatternFill("solid", fgColor=C_HEAD)
    for i, h in enumerate(_HEADERS, 1):
        c = ws.cell(HEAD_ROW, i, h)
        c.fill = head_fill
        c.font = Font(name="Helvetica", size=8, bold=True, color=WHITE)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    sum_lo, sum_hi = _col("Budget PM Hrs"), _col("Budget Manufacturing Hrs")
    col_sum, col_delta, col_total = _col("Σ Entered Hrs"), _col("Δ vs ETO Total"), _col("ETO Total Budget Hrs")

    for r_off, (_, s) in enumerate(seed_df.iterrows()):
        r = DATA_ROW + r_off
        pid = str(int(s["ProjectID"]))
        est_admin = float(s.get("EstAdminHours") or 0)
        est_eng   = float(s.get("EstEngHours") or 0)
        est_mfg   = float(s.get("EstMfgHours") or 0)
        est_total = float(s.get("LabEstHrs") or s.get("TotalBudgetHrs")
                          or (est_admin + est_eng + est_mfg))
        carried = prior.get(pid, {})

        # values, defaulting entry cells to carried-forward PM data or the ETO seed
        vals = {
            "Project ID": pid,
            "Project": s.get("Project", pid),
            "ETO Est Admin Hrs": round(est_admin, 1),
            "ETO Est Eng Hrs": round(est_eng, 1),
            "ETO Est Mfg Hrs": round(est_mfg, 1),
            "ETO Total Budget Hrs": round(est_total, 1),
            # pre-fill what ETO knows; leave Eng split blank for the PM
            "Budget PM Hrs": carried.get("Budget PM Hrs", round(est_admin, 1)),
            "Budget Mechanical Hrs": carried.get("Budget Mechanical Hrs"),
            "Budget Hydraulic Hrs": carried.get("Budget Hydraulic Hrs"),
            "Budget Electrical Hrs": carried.get("Budget Electrical Hrs"),
            "Budget Manufacturing Hrs": carried.get("Budget Manufacturing Hrs", round(est_mfg, 1)),
            "PO Ship Date": carried.get("PO Ship Date"),
            "Customer Agreed Ship Date": carried.get("Customer Agreed Ship Date"),
            "Planned Ship Date": carried.get("Planned Ship Date"),
            "% Done": carried.get("% Done"),
            "Material Budget $": carried.get("Material Budget $"),
        }
        for i, (h, role) in enumerate(_LAYOUT, 1):
            if h == "Σ Entered Hrs":
                v = f"=SUM({get_column_letter(sum_lo)}{r}:{get_column_letter(sum_hi)}{r})"
            elif h == "Δ vs ETO Total":
                v = f"={get_column_letter(col_sum)}{r}-{get_column_letter(col_total)}{r}"
            else:
                v = vals.get(h)
            c = ws.cell(r, i, v)
            c.border = border
            c.font = Font(name="Helvetica", size=9)
            fill = {"seed": C_SEED, "entry": C_ENTRY, "calc": C_CALC}[role]
            c.fill = PatternFill("solid", fgColor=fill)
            if "Hrs" in h or h in ("Σ Entered Hrs", "Δ vs ETO Total"):
                c.number_format = "#,##0"
            elif "Date" in h:
                c.number_format = "yyyy-mm-dd"
            elif h == "% Done":
                c.number_format = "0%"
            elif h == "Material Budget $":
                c.number_format = "$#,##0"

    # widths + freeze + legend
    for i, (h, _r) in enumerate(_LAYOUT, 1):
        ws.column_dimensions[get_column_letter(i)].width = 26 if h == "Project" else 13
    ws.freeze_panes = ws.cell(DATA_ROW, 3)
    note_r = DATA_ROW + len(seed_df) + 1
    ws.cell(note_r, 1,
            "Fill the yellow cells: split ETO's Engineering hours across Mechanical / "
            "Hydraulic / Electrical, adjust PM & Manufacturing if needed, and enter the "
            "ship dates. Keep 'Δ vs ETO Total' near 0.").font = \
        Font(name="Helvetica", size=8, italic=True, color=GREY)

    wb.save(output_path)
    return output_path


def _find_header_row(ws):
    for r in range(1, 6):
        if ws.cell(r, 1).value == "Project ID":
            return r
    return HEAD_ROW


def _read_entries(path):
    """Low-level: {ProjectID: {header: value}} of the entry cells (for carry-forward)."""
    try:
        wb = load_workbook(path, data_only=True)
    except Exception:
        return {}
    if BUDGETS_SHEET not in wb.sheetnames:
        return {}
    ws = wb[BUDGETS_SHEET]
    hr = _find_header_row(ws)
    headers = {ws.cell(hr, c).value: c for c in range(1, ws.max_column + 1)}
    if "Project ID" not in headers:
        return {}
    out = {}
    for r in range(hr + 1, ws.max_row + 1):
        pid = ws.cell(r, headers["Project ID"]).value
        if pid in (None, ""):
            continue
        pid = str(pid).split(".")[0].strip()
        out[pid] = {h: ws.cell(r, headers[h]).value for h in headers if h}
    return out


def read_budgets_input(path):
    """
    Read the Budgets (Input) sheet into an overlay-shaped DataFrame the dashboard
    consumes: ProjectID + BudgetHrs::<discipline> + ship dates + % done + material.
    Returns None if the sheet/columns aren't present.
    """
    entries = _read_entries(path)
    if not entries:
        return None
    rows = []
    for pid, cells in entries.items():
        row = {"ProjectID": pid}
        for header, key in _READ_MAP.items():
            row[key] = cells.get(header)
        rows.append(row)
    return pd.DataFrame(rows)


def seed_columns_present(seed_df):
    """Guard: seeding needs the ETO 3-bucket estimate hours on the summary frame."""
    need = {"ProjectID", "EstAdminHours", "EstEngHours", "EstMfgHours"}
    return need.issubset(set(seed_df.columns))
