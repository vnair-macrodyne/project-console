"""
console_diag_tabs.py — map the management pack's tabs (read-only, local file).

Dumps every sheet's name, dimensions, and first rows so the Budgets and PM Entries
tabs (the manual entry points) can be read the way they are ACTUALLY laid out —
rather than against an assumed schema. Also surfaces the Executive/Project Dashboard
and Data Validation structure.

Run:  python console_diag_tabs.py "Macrodyne Executive Dashboard.xlsx"
Paste the output back. Large data tabs (e.g. Labor Data) are capped to a few rows.
"""
import sys
from openpyxl import load_workbook

MAX_ROWS = 8       # header + a few sample rows per sheet
MAX_COLS = 40      # cap very wide sheets
FOCUS = ("budget", "pm entr", "pm_entr", "pmentr", "entries", "guide",
         "data validation", "dashboard", "worksheet", "calc")


def cellval(v):
    if v is None:
        return ""
    s = str(v).replace("\n", " ")
    return s[:22]


def dump_sheet(ws, deep):
    rows = min(ws.max_row, MAX_ROWS if not deep else max(MAX_ROWS, 14))
    cols = min(ws.max_column, MAX_COLS)
    print(f"\n{'='*90}\nSHEET: {ws.title}   ({ws.max_row} rows x {ws.max_column} cols)\n{'='*90}")
    if ws.max_column > MAX_COLS:
        print(f"  (showing first {MAX_COLS} of {ws.max_column} columns)")
    for r in range(1, rows + 1):
        vals = [cellval(ws.cell(r, c).value) for c in range(1, cols + 1)]
        # trim trailing empties for readability
        while vals and vals[-1] == "":
            vals.pop()
        if vals:
            print(f"  r{r:>2}: " + " | ".join(vals))


def main():
    if len(sys.argv) < 2:
        print('Usage: python console_diag_tabs.py "<workbook.xlsx>"')
        return
    wb = load_workbook(sys.argv[1], data_only=True, read_only=True)
    print("SHEETS:", wb.sheetnames)
    for name in wb.sheetnames:
        ws = wb[name]
        deep = any(k in name.lower() for k in FOCUS)
        dump_sheet(ws, deep)


if __name__ == "__main__":
    main()
