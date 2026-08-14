"""
etospec.py — the REAL eto-reporting report engine, vendored into the web console.

This is a faithful port of the deployed eto-reporting modules the user supplied as
the authoritative spec (eto_config / eto_daily / eto_excel / eto_postatus /
eto_exceptions).  The web Labour and Purchasing menus are driven straight from
this engine, and the Excel exports are produced by the same writers — so a report
opened or downloaded from the console is byte-for-byte the report eto-reporting
generates on the server.

What lives here:
  * The five DAILY labour reports  (A Departmental Project Detail, B Employee
    Summary, C Job-Category Summary, D Employee Job Detail, E Project Spend),
    each rendered in two views: "This Pay Period" and "Project Lifetime".
  * PO Status — On Order, Overdue   (Contents & Summary + PO Status sheets).
  * Procurement Exceptions by Buyer (flat, sortable, one row per project-item).

The SQL is the real string-contract SQL: base tables (tblTimecards / tblProjects /
tblCompany / tlkpHourTypes / tblEmployee) and the verified PO views.  Crucially,
Job Detail = tblTimecards.TimecardCustom1 (the field the deployed report reads),
NOT a spec/description column.
"""
import datetime as _dt

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io


COMPANY = "Macrodyne Technologies Inc."

# ─────────────────────────────────────────────────────────────────────────────
# Colours + row styling (verbatim from eto_config / eto_excel)
# ─────────────────────────────────────────────────────────────────────────────
COL_HEADER = "#1F3864"   # header background
COL_L1_SUB = "#D6E4F0"   # innermost subtotal (labour cat / employee subtotal)
COL_L2_SUB = "#BDD7EE"   # mid subtotal (project / employee)
COL_L3_SUB = "#9DC3E6"   # outer subtotal (project total / dept band)
COL_GRAND  = "#2E75B6"   # grand total — white text
COL_ALT    = "#F2F4F8"   # alternating detail row

# Row type → (fill ARGB, font ARGB, bold)
ROW_STYLES = {
    "l1_sub": ("FFD6E4F0", "FF000000", True),
    "l2_sub": ("FFBDD7EE", "FF000000", True),
    "l3_sub": ("FF9DC3E6", "FF000000", True),
    "grand":  ("FF2E75B6", "FFFFFFFF", True),
}

# Fields that render as plain integers (no decimals)
INTEGER_FIELDS = {"Employees", "UniqueEmps", "Entries"}


def hex_to_openpyxl(hex_str):
    return f"FF{hex_str.lstrip('#').upper()}"


def get_number_format(field_name, is_numeric):
    if field_name in INTEGER_FIELDS:
        return "0"
    if is_numeric:
        if "Cost" in field_name or "Payable" in field_name:
            return "$#,##0.00"
        return "#,##0.00"
    return "@"


def _populate_sheet(ws, title, subtitle, col_defs, row_tuples, period_label):
    """Populate a worksheet with title, subtitle, header, and styled data rows.
    Verbatim behaviour from eto_excel._populate_sheet."""
    ws["A1"] = title
    ws["A1"].font = Font(name="Helvetica", size=10, bold=True)
    ws.merge_cells(f"A1:{get_column_letter(len(col_defs))}1")

    ws["A2"] = f"{subtitle} — {period_label}"
    ws["A2"].font = Font(name="Helvetica", size=8, italic=True)
    ws.merge_cells(f"A2:{get_column_letter(len(col_defs))}2")

    data_start_row = 4
    ws.append([c[1] for c in col_defs])
    header_row = 3

    header_fill = PatternFill(start_color=hex_to_openpyxl(COL_HEADER),
                              end_color=hex_to_openpyxl(COL_HEADER), fill_type="solid")
    header_font = Font(name="Helvetica", size=7, bold=True, color="FFFFFFFF")
    for col_idx, _ in enumerate(col_defs, 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    detail_count = 0
    thin_border = Border(left=Side(style="thin", color="CCCCCC"),
                         right=Side(style="thin", color="CCCCCC"),
                         top=Side(style="thin", color="CCCCCC"),
                         bottom=Side(style="thin", color="CCCCCC"))

    for row_idx, (cells, row_type) in enumerate(row_tuples, data_start_row):
        ws.append(list(cells))
        for col_idx, (col_def, cell_val) in enumerate(zip(col_defs, cells), 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            field_name, is_numeric, align = col_def[0], col_def[4], col_def[3]
            fmt = get_number_format(field_name, is_numeric)
            if fmt != "@":
                cell.number_format = fmt
            if align == "R":
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif align == "C":
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = thin_border
            if row_type == "detail":
                if detail_count % 2 == 1:
                    cell.fill = PatternFill(start_color=hex_to_openpyxl(COL_ALT),
                                            end_color=hex_to_openpyxl(COL_ALT), fill_type="solid")
                detail_count += 1
            elif row_type in ROW_STYLES:
                color_argb, text_color_argb, bold = ROW_STYLES[row_type]
                cell.fill = PatternFill(start_color=color_argb, end_color=color_argb, fill_type="solid")
                cell.font = Font(name="Helvetica", size=7, bold=bold, color=text_color_argb)

    for col_idx, col_def in enumerate(col_defs, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_def[2] * 3.5
    ws.freeze_panes = f"A{data_start_row}"


# ─────────────────────────────────────────────────────────────────────────────
# Pay-period helpers (verbatim from eto_paths)
# ─────────────────────────────────────────────────────────────────────────────
BIWEEKLY_ANCHOR = _dt.date(2026, 7, 6)   # a known pay-period START (Monday)
PERIOD_LEN = 14


def biweekly_period(target=None):
    target = target or _dt.date.today()
    n = (target - BIWEEKLY_ANCHOR).days // PERIOD_LEN
    start = BIWEEKLY_ANCHOR + _dt.timedelta(days=n * PERIOD_LEN)
    return start, start + _dt.timedelta(days=PERIOD_LEN - 1)


def period_to_date(as_of=None):
    as_of = as_of or _dt.date.today()
    start, _ = biweekly_period(as_of)
    return start, as_of


# ─────────────────────────────────────────────────────────────────────────────
# LABOUR — column defs + query + builders (verbatim from eto_daily)
# ─────────────────────────────────────────────────────────────────────────────
COLS_A = [
    ("Department", "Department", 12, "L", False),
    ("ProjectID",  "Project ID", 7,  "L", False),
    ("JobName",    "Job Name",   22, "L", False),
    ("Customer",   "Customer",   16, "L", False),
    ("EmpNo",      "Emp No",     6,  "L", False),
    ("Employee",   "Employee",   14, "L", False),
    ("Entries",    "Entries",    6,  "R", True),
    ("Hours",      "Hours",      7,  "R", True),
    ("OTHours",    "OT Hours",   7,  "R", True),
    ("LabourCost", "Labour Cost", 10, "R", True),
]
COLS_B = [
    ("Department", "Department", 14, "L", False),
    ("EmpNo",      "Emp No",     7,  "L", False),
    ("Employee",   "Employee",   18, "L", False),
    ("Entries",    "Entries",    7,  "R", True),
    ("Hours",      "Hours",      8,  "R", True),
    ("OTHours",    "OT Hours",   8,  "R", True),
    ("LabourCost", "Labour Cost", 11, "R", True),
]
COLS_C = [
    ("Department", "Department",      14, "L", False),
    ("Category",   "Labour Category", 20, "L", False),
    ("Entries",    "Entries",         7,  "R", True),
    ("Hours",      "Hours",           8,  "R", True),
    ("OTHours",    "OT Hours",        8,  "R", True),
    ("LabourCost", "Labour Cost",     11, "R", True),
    ("PctDept",    "% of Dept Hrs",   9,  "R", True),
]
COLS_D = [
    ("ProjectID",  "Project",         7,  "L", False),
    ("JobName",    "Job Name",        20, "L", False),
    ("Category",   "Labour Category", 16, "L", False),
    ("JobDetail",  "Job Detail",      18, "L", False),
    ("Hours",      "Hours",           7,  "R", True),
    ("OTHours",    "OT Hours",        7,  "R", True),
    ("LabourCost", "Labour Cost",     11, "R", True),
]
COLS_E = [
    ("ProjectID",  "Project ID",  7,  "L", False),
    ("JobName",    "Job Name",    22, "L", False),
    ("Customer",   "Customer",    16, "L", False),
    ("Department", "Department",  12, "L", False),
    ("EmpNo",      "Emp No",      6,  "L", False),
    ("Employee",   "Employee",    14, "L", False),
    ("Entries",    "Entries",     6,  "R", True),
    ("Hours",      "Hours",       7,  "R", True),
    ("OTHours",    "OT Hours",    7,  "R", True),
    ("LabourCost", "Labour Cost", 10, "R", True),
]

_AGG = ["Entries", "Hours", "OTHours", "LabourCost"]


def query_daily_labour(start_date, end_date, project_ids=None):
    """Granular labour rows for TimeDate up to end-of-day(end). Half-open upper bound.
    start_date=None -> no lower bound (project lifetime-to-date). Optional project scope.
    Job Detail = tblTimecards.TimecardCustom1 (the deployed report's source field)."""
    end_next = end_date + _dt.timedelta(days=1)
    lower = f"tc.TimeDate >= '{start_date:%Y-%m-%d}' AND " if start_date else ""
    proj = ""
    if project_ids:
        ids = ",".join(str(int(p)) for p in project_ids)
        proj = f" AND tc.ProjectID IN ({ids})"
    return f"""
    SELECT
        ht.HourDepartment                           AS Department,
        tc.ProjectID                                AS ProjectID,
        p.DisplayName                               AS JobName,
        cust.CName                                  AS Customer,
        tc.EmpNumber                                AS EmpNo,
        tc.EmpNumber + ' - ' + e.EmpLastName + ', ' + e.EmpFirstName AS Employee,
        CASE WHEN tc.SpecID = 999 THEN 'Re-work'
             ELSE ht.HourDescription END             AS Category,
        tc.SpecID                                   AS Machine,
        ISNULL(tc.TimecardCustom1, '')              AS JobDetail,
        1                                           AS Entries,
        tc.HourTime                                 AS Hours,
        CASE WHEN tc.HourFactor > 1 THEN tc.HourTime ELSE 0 END AS OTHours,
        tc.HourTime * tc.HourRate * tc.HourFactor   AS LabourCost
    FROM tblTimecards tc
    JOIN      tblProjects  p    ON p.ProjectID   = tc.ProjectID
    LEFT JOIN tblCompany   cust ON cust.CompanyID = p.CompanyID
    LEFT JOIN tlkpHourTypes ht  ON ht.HourType   = tc.HourType
    LEFT JOIN tblEmployee  e    ON e.EmployeeID  = tc.EmployeeID
    WHERE {lower}tc.TimeDate < '{end_next:%Y-%m-%d}'{proj}
    """


def _sums(df):
    return [int(df["Entries"].sum()), round(float(df["Hours"].sum()), 2),
            round(float(df["OTHours"].sum()), 2), round(float(df["LabourCost"].sum()), 2)]


def build_report_a(df):
    """Department -> Project -> Employee (one row per employee per project)."""
    rows = []
    for dept in sorted(df["Department"].dropna().unique()):
        dsub = df[df["Department"] == dept]
        for pid in sorted(dsub["ProjectID"].dropna().unique(), key=str):
            psub = dsub[dsub["ProjectID"] == pid]
            job, cust = psub["JobName"].iloc[0], psub["Customer"].iloc[0]
            g = (psub.groupby(["EmpNo", "Employee"], dropna=False)[_AGG]
                     .sum().reset_index().sort_values("Employee"))
            for _, r in g.iterrows():
                rows.append(([dept, pid, job, cust, r.EmpNo, r.Employee,
                              int(r.Entries), round(r.Hours, 2), round(r.OTHours, 2),
                              round(r.LabourCost, 2)], "detail"))
            e, h, ot, c = _sums(psub)
            rows.append((["", pid, f"{job} — Project Subtotal", "", "", "", e, h, ot, c], "l1_sub"))
        e, h, ot, c = _sums(dsub)
        rows.append(([f"{dept} — Department Total", "", "", "", "", "", e, h, ot, c], "l2_sub"))
    e, h, ot, c = _sums(df)
    rows.append((["GRAND TOTAL", "", "", "", "", "", e, h, ot, c], "grand"))
    return rows


def build_report_b(df):
    """Department -> Employee (one row per employee, summed across projects)."""
    rows = []
    for dept in sorted(df["Department"].dropna().unique()):
        dsub = df[df["Department"] == dept]
        g = (dsub.groupby(["EmpNo", "Employee"], dropna=False)[_AGG]
                 .sum().reset_index().sort_values("Employee"))
        for _, r in g.iterrows():
            rows.append(([dept, r.EmpNo, r.Employee, int(r.Entries),
                          round(r.Hours, 2), round(r.OTHours, 2), round(r.LabourCost, 2)], "detail"))
        e, h, ot, c = _sums(dsub)
        rows.append(([f"{dept} — Department Total", "", "", e, h, ot, c], "l1_sub"))
    e, h, ot, c = _sums(df)
    rows.append((["GRAND TOTAL", "", "", e, h, ot, c], "grand"))
    return rows


def build_report_c(df):
    """Department -> Category (HourDescription), with % of dept hours."""
    rows = []
    for dept in sorted(df["Department"].dropna().unique()):
        dsub = df[df["Department"] == dept]
        dept_hours = float(dsub["Hours"].sum()) or 1.0
        g = (dsub.groupby(["Category"], dropna=False)[_AGG]
                 .sum().reset_index().sort_values("Hours", ascending=False))
        for _, r in g.iterrows():
            pct = round(100.0 * float(r.Hours) / dept_hours, 1)
            rows.append(([dept, r.Category, int(r.Entries), round(r.Hours, 2),
                          round(r.OTHours, 2), round(r.LabourCost, 2), pct], "detail"))
        e, h, ot, c = _sums(dsub)
        rows.append(([f"{dept} — Department Total", "", e, h, ot, c, 100.0], "l1_sub"))
    e, h, ot, c = _sums(df)
    rows.append((["GRAND TOTAL", "", e, h, ot, c, ""], "grand"))
    return rows


def build_report_d(df):
    """Department -> Employee -> LABOUR CATEGORY -> job-detail rows (cumulative). The job details
    are grouped by labour category with a per-category subtotal, then an employee subtotal and a
    department total."""
    rows = []
    n = len(COLS_D)
    band = lambda label, rt: rows.append(([label] + [""] * (n - 1), rt))
    for dept in sorted(df["Department"].dropna().unique()):
        dsub = df[df["Department"] == dept]
        band(f"Department: {dept}", "l3_sub")
        emps = dsub[["EmpNo", "Employee"]].drop_duplicates().sort_values("Employee")
        for _, er in emps.iterrows():
            esub = dsub[dsub["EmpNo"] == er.EmpNo]
            band(f"{er.Employee}", "l2_sub")
            for cat in sorted(esub["Category"].dropna().unique()):
                csub = esub[esub["Category"] == cat]
                g = (csub.groupby(["ProjectID", "JobName", "JobDetail"], dropna=False)[_AGG]
                         .sum().reset_index().sort_values(["ProjectID", "JobDetail"]))
                for _, r in g.iterrows():
                    rows.append(([r.ProjectID, r.JobName, cat, r.JobDetail,
                                  round(r.Hours, 2), round(r.OTHours, 2), round(r.LabourCost, 2)],
                                 "detail"))
                _, h, ot, c = _sums(csub)
                rows.append((["", "", f"{cat} — subtotal", "", h, ot, c], "l1_sub"))
            _, h, ot, c = _sums(esub)
            rows.append(([f"{er.Employee} — Subtotal", "", "", "", h, ot, c], "l2_sub"))
        _, h, ot, c = _sums(dsub)
        rows.append(([f"{dept} — Department Total", "", "", "", h, ot, c], "l3_sub"))
    _, h, ot, c = _sums(df)
    rows.append((["GRAND TOTAL", "", "", "", h, ot, c], "grand"))
    return rows


def build_report_e(df):
    """Project Labour Spend: Project -> Department -> Employee (cumulative)."""
    rows = []
    for pid in sorted(df["ProjectID"].dropna().unique(), key=str):
        psub = df[df["ProjectID"] == pid]
        job, cust = psub["JobName"].iloc[0], psub["Customer"].iloc[0]
        for dept in sorted(psub["Department"].dropna().unique()):
            dsub = psub[psub["Department"] == dept]
            g = (dsub.groupby(["EmpNo", "Employee"], dropna=False)[_AGG]
                     .sum().reset_index().sort_values("Employee"))
            for _, r in g.iterrows():
                rows.append(([pid, job, cust, dept, r.EmpNo, r.Employee,
                              int(r.Entries), round(r.Hours, 2), round(r.OTHours, 2),
                              round(r.LabourCost, 2)], "detail"))
            e, h, ot, c = _sums(dsub)
            rows.append((["", "", "", f"{dept} — Dept Subtotal", "", "", e, h, ot, c], "l1_sub"))
        e, h, ot, c = _sums(psub)
        rows.append(([pid, f"{job} — Project Total", "", "", "", "", e, h, ot, c], "l2_sub"))
    e, h, ot, c = _sums(df)
    rows.append((["GRAND TOTAL", "", "", "", "", "", e, h, ot, c], "grand"))
    return rows


COLS_DISC = [
    ("ProjectID",  "Project ID",  8,  "L", False),
    ("JobName",    "Job Name",    22, "L", False),
    ("Discipline", "Discipline",  18, "L", False),
    ("Employees",  "Emps",        6,  "R", True),
    ("Entries",    "Entries",     7,  "R", True),
    ("Hours",      "Hours",       8,  "R", True),
    ("OTHours",    "OT Hours",    8,  "R", True),
    ("LabourCost", "Labour Cost", 11, "R", True),
]

# canonical discipline order for the by-discipline report (matches the dashboard crosswalk)
_DISC_ORDER = ["Project Management", "Mechanical Engineering", "Electrical Engineering",
               "Hydraulic Engineering", "Manufacturing", "Other", "Re-work"]


def _disc_sort_key(d):
    ds = "" if d is None else str(d)
    return (_DISC_ORDER.index(ds), "") if ds in _DISC_ORDER else (len(_DISC_ORDER), ds)


def build_report_disc(df):
    """Project -> Discipline (the 6-discipline crosswalk + a Re-work bucket): headcount, hours,
    OT and applied-rate cost per discipline, with a per-project subtotal and a grand total.
    Requires a 'Discipline' column on df (added by the query layer via the crosswalk)."""
    rows = []
    for pid in sorted(df["ProjectID"].dropna().unique(), key=str):
        psub = df[df["ProjectID"] == pid]
        job = psub["JobName"].iloc[0]
        gsum = psub.groupby("Discipline", dropna=False)[_AGG].sum()
        gemp = psub.groupby("Discipline", dropna=False)["EmpNo"].nunique()
        for disc in sorted(gsum.index, key=_disc_sort_key):
            a = gsum.loc[disc]
            label = "(unclassified)" if disc is None or (isinstance(disc, float)) else str(disc)
            rows.append(([pid, job, label, int(gemp.loc[disc]), int(a.Entries),
                          round(float(a.Hours), 2), round(float(a.OTHours), 2),
                          round(float(a.LabourCost), 2)], "detail"))
        e, h, ot, c = _sums(psub)
        rows.append(([f"Project {pid} — total", "", "", int(psub["EmpNo"].nunique()),
                      e, h, ot, c], "l1_sub"))
    e, h, ot, c = _sums(df)
    rows.append((["GRAND TOTAL", "", "", int(df["EmpNo"].nunique()), e, h, ot, c], "grand"))
    return rows


COLS_DSUM = [
    ("Machine",    "Machine",     8,  "L", False),
    ("JobDetail",  "Job Detail",  32, "L", False),
    ("Entries",    "Entries",     7,  "R", True),
    ("Hours",      "Hours",       8,  "R", True),
    ("OTHours",    "OT Hours",    8,  "R", True),
    ("LabourCost", "Labour Cost", 11, "R", True),
]


def _mach_label(v):
    """Timecard SpecID -> machine number ('10.0' -> '10'); blank when the charge has no machine."""
    f = _num_or_none(v)
    if f is None:
        return ""
    return str(int(f)) if float(f).is_integer() else str(f)


def _mach_sort(m):
    """Numeric machines ascending; the '(no machine)' bucket last."""
    if m == "":
        return (1, 0.0, "")
    try:
        return (0, float(m), "")
    except (TypeError, ValueError):
        return (0, 0.0, m)


def build_report_dsum(df):
    """Project -> MACHINE number -> job-detail SUMMARY (cumulative). Job details are summed across
    everything else (employee, category) — one row per (machine, job detail) — with a per-machine
    subtotal and a project total. Machine = the timecard SpecID; charges with no machine fall in a
    '(no machine)' bucket."""
    import pandas as pd
    rows = []
    n = len(COLS_DSUM)
    band = lambda label, rt: rows.append(([label] + [""] * (n - 1), rt))
    d = df.copy()
    d["_M"] = d["Machine"].apply(_mach_label) if "Machine" in d.columns else ""
    for pid in sorted(d["ProjectID"].dropna().unique(), key=str):
        psub = d[d["ProjectID"] == pid]
        job = psub["JobName"].iloc[0]
        band(f"Project: {pid} — {job}", "l3_sub")
        for mach in sorted(psub["_M"].unique(), key=_mach_sort):
            msub = psub[psub["_M"] == mach]
            mlabel = mach if mach else "(no machine)"
            g = (msub.groupby(["JobDetail"], dropna=False)[_AGG]
                     .sum().reset_index().sort_values("JobDetail"))
            for _, r in g.iterrows():
                jd = r.JobDetail
                blank = jd is None or (isinstance(jd, float) and pd.isna(jd)) or str(jd).strip() == ""
                jd = "(no detail)" if blank else str(jd)
                rows.append(([mlabel, jd, int(r.Entries), round(r.Hours, 2),
                              round(r.OTHours, 2), round(r.LabourCost, 2)], "detail"))
            e, h, ot, c = _sums(msub)
            rows.append(([f"Machine {mlabel} — subtotal", "", e, h, ot, c], "l1_sub"))
        e, h, ot, c = _sums(psub)
        rows.append(([f"Project {pid} — Total", "", e, h, ot, c], "l2_sub"))
    e, h, ot, c = _sums(df)
    rows.append((["GRAND TOTAL", "", e, h, ot, c], "grand"))
    return rows


# report registry: id -> metadata + column defs + builder
LABOUR_REPORTS = {
    "lab_a": dict(order=0, label="Departmental Project Detail",
                  title="ETO Daily Departmental Project Detail",
                  suffix="A_Departmental_Project_Detail", cols=COLS_A, builder=build_report_a),
    "lab_b": dict(order=1, label="Employee Summary",
                  title="ETO Daily Employee Summary",
                  suffix="B_Employee_Summary", cols=COLS_B, builder=build_report_b),
    "lab_c": dict(order=2, label="Job-Category Summary",
                  title="ETO Daily Job-Category Summary",
                  suffix="C_Job_Category_Summary", cols=COLS_C, builder=build_report_c),
    "lab_d": dict(order=3, label="Employee Job Detail",
                  title="ETO Daily Employee Job Detail",
                  suffix="D_Employee_Job_Detail", cols=COLS_D, builder=build_report_d),
    "lab_e": dict(order=4, label="Project Labour Spend",
                  title="ETO Daily Project Labour Spend",
                  suffix="E_Project_Spend", cols=COLS_E, builder=build_report_e),
    "lab_disc": dict(order=5, label="By Discipline",
                     title="ETO Labour by Discipline",
                     suffix="F_By_Discipline", cols=COLS_DISC, builder=build_report_disc),
    "lab_dsum": dict(order=6, label="Job Detail Summary",
                     title="ETO Daily Job Detail Summary — by Project & Machine",
                     suffix="G_Job_Detail_Summary", cols=COLS_DSUM, builder=build_report_dsum),
}


def labour_book_bytes(report_id, grouped_rows, label):
    """Single-sheet workbook for the selected time window (cumulative to the end date)."""
    meta = LABOUR_REPORTS[report_id]
    wb = Workbook(); wb.remove(wb.active)
    ws = wb.create_sheet(title=(meta["label"][:31] or "Labour"))
    _populate_sheet(ws, meta["title"], "Applied-rate labour spend", meta["cols"], grouped_rows, label)
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# PO STATUS — On Order, Overdue (verbatim from eto_postatus)
# ─────────────────────────────────────────────────────────────────────────────
COLS_PO = [
    ("Item",        "Item",        16, "L", False),
    ("Description", "Description", 26, "L", False),
    ("PO",          "PO #",         7, "L", False),
    ("Supplier",    "Supplier",    18, "L", False),
    ("Qty",         "Qty",          6, "R", True),
    ("Received",    "Rec'd",        6, "R", True),
    ("Price",       "Price",        9, "R", True),
    ("ExtValue",    "Ext. Value",  10, "R", True),
    ("Required",    "Required",    10, "C", False),
    ("Revised",     "Revised",     10, "C", False),
    ("DaysLate",    "Days Late",    8, "R", True),
    ("Status",      "Status",       8, "L", False),
]

AGE_BUCKETS = [("<=30 days", 0, 30), ("31-90 days", 31, 90),
               ("91-365 days", 91, 365), (">365 days (stale)", 366, 10 ** 9)]


def _po_date_window(date_from=None, date_to=None):
    """Optional filter on the PO-placed date (poh.PurchaseDate). A missing bound = open-ended,
    so 'From Project Start' (date_from=None) keeps the lower side open, like the labour reports."""
    parts = []
    if date_from:
        parts.append(f"CAST(poh.PurchaseDate AS date) >= '{date_from}'")
    if date_to:
        parts.append(f"CAST(poh.PurchaseDate AS date) <= '{date_to}'")
    return (" AND " + " AND ".join(parts)) if parts else ""


def query_po_status_open(project_ids=None, date_from=None, date_to=None):
    proj = ""
    if project_ids:
        ids = ",".join(str(int(p)) for p in project_ids)
        proj = f" AND pod.ProjectID IN ({ids})"
    dt = _po_date_window(date_from, date_to)
    return f"""
    SELECT
        pod.ProjectID                 AS ProjectID,
        p.DisplayName                 AS JobName,
        pcust.CName                   AS Customer,
        pod.SpecID                    AS MachineCode,
        pod.ItemID                    AS Item,
        pod.ItemDescription           AS Description,
        poh.PurchaseOrderID           AS PO,
        poh.CName                     AS Supplier,
        COALESCE(bu.EmpLastName + ', ' + bu.EmpFirstName,
                 CAST(poh.BuyerID AS varchar(20)))  AS Buyer,
        p.PStatus                     AS ProjStatus,
        pod.PurchaseQty               AS Qty,
        pod.Received                  AS Received,
        pod.PurchasePrice             AS Price,
        pod.ExtendedPrice             AS ExtValue,
        CAST(pod.DateRequired AS date) AS Required,
        CAST(pod.DateRevised  AS date) AS Revised
    FROM vwPurchaseOrderHeader poh
    JOIN vwPurchaseOrderDetails pod ON pod.PurchaseOrderID = poh.PurchaseOrderID
    LEFT JOIN tblProjects p     ON p.ProjectID  = pod.ProjectID
    LEFT JOIN tblCompany  pcust ON pcust.CompanyID = p.CompanyID
    LEFT JOIN tblEmployee bu    ON bu.EmployeeID = poh.BuyerID
    WHERE poh.PurchaseActive = 1
      AND (pod.Received IS NULL OR pod.Received < pod.PurchaseQty){proj}{dt}
    ORDER BY pod.ProjectID, pod.SpecID, poh.PurchaseOrderID, pod.ItemID
    """


def po_prep(df, today=None):
    import pandas as pd
    today = today or _dt.date.today()
    df = df.copy()
    req = pd.to_datetime(df["Required"], errors="coerce")
    rev = pd.to_datetime(df["Revised"], errors="coerce")
    oper = rev.fillna(req)
    df["Required"] = [d.date() if pd.notna(d) else None for d in req]
    df["Revised"] = [d.date() if pd.notna(d) else None for d in rev]
    df["DaysLate"] = [(today - d.date()).days if (pd.notna(d) and d.date() < today) else 0 for d in oper]
    df["Status"] = ["Overdue" if dl > 0 else "On Order" for dl in df["DaysLate"]]
    return df


def po_build_rows(df):
    rows, proj_index = [], []
    for pid in sorted(df["ProjectID"].dropna().unique(), key=str):
        psub = df[df["ProjectID"] == pid]
        job, cust = psub["JobName"].iloc[0], psub["Customer"].iloc[0]
        pstat = psub["ProjStatus"].iloc[0] if "ProjStatus" in psub.columns else None
        flag = "" if (pstat == "Sold") else f"   [!] NON-ACTIVE PROJECT ({pstat or 'unknown'}) — review/cancel"
        proj_index.append((pid, len(rows)))
        rows.append(([f"Project: {pid} — {job}   ·   Customer: {cust}{flag}"] + [""] * 10, "l3_sub"))
        for mc in sorted(psub["MachineCode"].dropna().unique(), key=str):
            msub = psub[psub["MachineCode"] == mc]
            rows.append(([f"Machine {mc}"] + [""] * 10, "l2_sub"))
            for _, r in msub.iterrows():
                rows.append(([r.Item, r.Description, r.PO, r.Supplier, r.Qty, r.Received,
                              round(r.Price, 2), round(r.ExtValue, 2),
                              str(r.Required or ""), str(r.Revised or ""),
                              (int(r.DaysLate) if r.Status == "Overdue" else ""), r.Status], "detail"))
        rows.append((["", f"Project {pid} — Open Value", "", "", "", "", "",
                      round(psub["ExtValue"].sum(), 2), "", "", "", ""], "l1_sub"))
    rows.append((["GRAND TOTAL — Open PO Value"] + [""] * 6 + [round(df["ExtValue"].sum(), 2), "", "", "", ""], "grand"))
    return rows, proj_index


def po_aging_summary(df):
    ov = df[df["Status"] == "Overdue"]
    out = []
    for name, lo, hi in AGE_BUCKETS:
        b = ov[(ov["DaysLate"] >= lo) & (ov["DaysLate"] <= hi)]
        out.append((name, len(b), round(float(b["ExtValue"].sum()), 2)))
    return out


def _po_write_contents(ws, df, label, proj_index, detail_sheet="PO Status", data_start=4):
    NAVY = "1F3864"; hfont = Font(bold=True, color="FFFFFF"); hfill = PatternFill("solid", fgColor=NAVY)
    ws["A1"] = "PO Status — Table of Contents & Summary"; ws["A1"].font = Font(bold=True, size=12)
    ws["A2"] = label; ws["A2"].font = Font(italic=True, size=9)
    r = 4
    ws.cell(r, 1, "Summary — Overdue Aging").font = Font(bold=True, size=10); r += 1
    for i, h in enumerate(["Age Bucket", "Lines", "Ext. Value"], 1):
        c = ws.cell(r, i, h); c.font = hfont; c.fill = hfill
    r += 1
    tl = tv = 0
    for name, lines, val in po_aging_summary(df):
        ws.cell(r, 1, name); ws.cell(r, 2, lines); ws.cell(r, 3, val).number_format = "$#,##0.00"
        tl += lines; tv += val; r += 1
    ws.cell(r, 1, "Total Overdue").font = Font(bold=True)
    ws.cell(r, 2, tl).font = Font(bold=True)
    ws.cell(r, 3, round(tv, 2)).font = Font(bold=True); ws.cell(r, 3).number_format = "$#,##0.00"; r += 1
    if "ProjStatus" in df.columns:
        na = df[(df["Status"] == "Overdue") & (df["ProjStatus"] != "Sold")]
        ws.cell(r, 1, "— of which on non-active projects").font = Font(italic=True, color="C00000")
        ws.cell(r, 2, len(na)).font = Font(italic=True, color="C00000")
        cc = ws.cell(r, 3, round(float(na["ExtValue"].sum()), 2)); cc.font = Font(italic=True, color="C00000")
        cc.number_format = "$#,##0.00"; r += 1
    r += 1
    ws.cell(r, 1, "Contents — Projects").font = Font(bold=True, size=10); r += 1
    for i, h in enumerate(["Project", "Status", "Open Lines", "Open Value", "Overdue Value"], 1):
        c = ws.cell(r, i, h); c.font = hfont; c.fill = hfill
    r += 1
    off_by_pid = dict(proj_index)
    for pid in sorted(df["ProjectID"].dropna().unique(), key=str):
        psub = df[df["ProjectID"] == pid]
        pstat = psub["ProjStatus"].iloc[0] if "ProjStatus" in psub.columns else None
        ov = psub[psub["Status"] == "Overdue"]
        cell = ws.cell(r, 1, str(pid))
        if pid in off_by_pid:
            cell.hyperlink = f"#'{detail_sheet}'!A{data_start + off_by_pid[pid]}"
            cell.font = Font(color="1F3864", underline="single")
        ws.cell(r, 2, pstat or "unknown")
        if pstat != "Sold":
            ws.cell(r, 2).font = Font(color="C00000", bold=True)
        ws.cell(r, 3, int(len(psub)))
        ws.cell(r, 4, round(float(psub["ExtValue"].sum()), 2)).number_format = "$#,##0.00"
        ws.cell(r, 5, round(float(ov["ExtValue"].sum()), 2)).number_format = "$#,##0.00"
        r += 1
    for col, w in zip("ABCDE", (30, 12, 11, 16, 16)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A3"


def po_status_book_bytes(df, label):
    """Full workbook: Contents & Summary (landing) + PO Status detail. df must be po_prep'd."""
    wb = Workbook(); wb.remove(wb.active)
    rows, proj_index = po_build_rows(df)
    ws = wb.create_sheet("PO Status", 1)
    _populate_sheet(ws, "PO Status — On Order, Overdue", "Open purchase order lines",
                    COLS_PO, rows, label)
    for fld in ("Price", "ExtValue"):
        ci = [i for i, c in enumerate(COLS_PO, 1) if c[0] == fld][0]
        for row in range(4, ws.max_row + 1):
            ws.cell(row=row, column=ci).number_format = "$#,##0.00"
    _po_write_contents(wb.create_sheet("Contents & Summary", 0), df, label, proj_index)
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# PROCUREMENT EXCEPTIONS by Buyer (verbatim from eto_exceptions)
# ─────────────────────────────────────────────────────────────────────────────
COLS_FLAT = [
    ("Buyer",       "Buyer",       16, "L", False),
    ("ProjectID",   "Project #",    8, "C", False),
    ("JobName",     "Project",     20, "L", False),
    ("Item",        "Item",        14, "L", False),
    ("Description", "Description", 24, "L", False),
    ("POs",         "PO #(s)",     16, "L", False),
    ("Vendors",     "Vendor(s)",   18, "L", False),
    ("POCount",     "#POs",         5, "R", True),
    ("ExtValue",    "Ext. Value",  11, "R", True),
    ("NeedBy",      "Need-By",     10, "C", False),
    ("DaysLate",    "Days Late",    7, "R", True),
    ("EngRelease",  "Eng Rel'd",   10, "C", False),
    ("LLT",         "LLT",          5, "C", False),
    ("Oversized",   "Oversize",     8, "C", False),
    ("DelLate",     "Del Late",     7, "C", False),
]

# custom-field layout (confirmed via tlkpCaption 2026-07-22)
_LLT_FLAG_COL = "PartCustom7"; _OVERSIZE_FLAG_COL = "PartCustom8"; _ENG_RELEASE_COL = "PartCustom6"
_LLT_CRITICAL_DAYS = 90

# ── PO Exceptions: PER-LINE detail layout (the 19 fields Vijay requested, 2026-08-12) ──────────
# Populated from ETO: Buyer, Project#, Project, Code(spec), Item, Category, Release date, PO#,
# Planned/Revised/Receipt dates, Status(derived), Lead time, Oversized. BLANK — ETO holds no source:
# Planned Ship (BudgetShipRelease empty), Days to Assembly (MfgBegin empty), RFQ Date, Permit Dates,
# Last Updated (no PO-line modified timestamp). Kept as columns so the layout matches the workbook.
COLS_EXC = [
    ("Buyer",          "Buyer",            16, "L", False),
    ("ProjectID",      "Project #",         8, "C", False),
    ("JobName",        "Project",          20, "L", False),
    ("Code",           "Code",              6, "C", False),
    ("Item",           "Item",             11, "L", False),
    ("Category",       "Category",         16, "L", False),
    ("EngRelease",     "Release Date",     11, "C", False),
    ("PO",             "PO #",             10, "L", False),
    ("PlannedShip",    "Planned Ship",     11, "C", False),
    ("PlannedReceipt", "Planned Receipt",  12, "C", False),
    ("RevisedReceipt", "Revised Receipt",  12, "C", False),
    ("ReceiptDate",    "Receipt Date",     11, "C", False),
    ("Status",         "Status",           12, "L", False),
    ("LastUpdated",    "Last Updated",     11, "C", False),
    ("DaysToAssembly", "Days to Assembly",  8, "R", True),
    ("RFQDate",        "RFQ Date",         11, "C", False),
    ("PermitDates",    "Permit Dates",     12, "L", False),
    ("LeadTime",       "Lead Time",         8, "R", True),
    ("Oversized",      "Oversized",         8, "C", False),
]


def _num_or_none(v):
    import pandas as pd
    try:
        f = float(v)
        return None if pd.isna(f) else f
    except (TypeError, ValueError):
        return None


def _spec_code(v):
    """SpecID 10.0 -> '10' (machine/spec code); blank-safe."""
    f = _num_or_none(v)
    if f is None:
        return ""
    return str(int(f)) if float(f).is_integer() else str(f)


def exc_detail(df, today=None):
    """One normalized row per OPEN, OVERDUE PO line — the full field set for the detail report.
    Exceptions = open lines (not fully received) whose need-by (revised else required) has passed."""
    import pandas as pd
    today = today or _dt.date.today()
    out = []
    for _, r in df.iterrows():
        req = pd.to_datetime(r.get("DateRequired"), errors="coerce")
        rev = pd.to_datetime(r.get("DateRevised"), errors="coerce")
        need = rev if pd.notna(rev) else req
        needd = need.date() if pd.notna(need) else None
        if not (needd and needd < today):                 # exceptions = OVERDUE open lines
            continue
        rcpt = pd.to_datetime(r.get("ReceiptDate"), errors="coerce")
        eng = pd.to_datetime(r.get("EngReleaseDate"), errors="coerce")
        recv = _num_or_none(r.get("Received")) or 0.0
        status = "Overdue — partial" if recv > 0 else "Overdue"
        pid = r.get("ProjectID")
        out.append({
            "Buyer": (str(r.get("Buyer")).strip() if not _blank(r.get("Buyer")) else "(unassigned)"),
            "ProjectID": ("" if _blank(pid) else str(int(pid)) if _num_or_none(pid)
                          and float(pid).is_integer() else str(pid)),
            "JobName": ("" if _blank(r.get("JobName")) else str(r.get("JobName"))),
            "Code": _spec_code(r.get("Code")),
            "Item": ("" if _blank(r.get("Item")) else str(r.get("Item"))),
            "Category": ("" if _blank(r.get("Category")) else str(r.get("Category"))),
            "EngRelease": (eng.date().isoformat() if pd.notna(eng) else ""),
            "PO": ("" if _blank(r.get("PO")) else str(r.get("PO"))),
            "PlannedShip": "",                            # ETO holds no maintained ship date
            "PlannedReceipt": (req.date().isoformat() if pd.notna(req) else ""),
            "RevisedReceipt": (rev.date().isoformat() if pd.notna(rev) else ""),
            "ReceiptDate": (rcpt.date().isoformat() if pd.notna(rcpt) else ""),
            "Status": status,
            "LastUpdated": "",                            # no PO-line modified timestamp in ETO
            "DaysToAssembly": None,                       # no maintained assembly date
            "RFQDate": "",                                # not in ETO
            "PermitDates": "",                            # not in ETO
            "LeadTime": (int(_num_or_none(r.get("LeadDays"))) if _num_or_none(r.get("LeadDays")) else None),
            "Oversized": ("yes" if _flag(r.get("OverFlag")) else ""),
            "ExtValue": round(_num_or_none(r.get("ExtValue")) or 0.0, 2),
            "DaysLate": (today - needd).days,
        })
    return pd.DataFrame(out)


def exc_detail_build_rows(items):
    """(cells, kind) rows for the per-line detail report — sorted by buyer then most-overdue."""
    if items is None or items.empty:
        return [(["No open-PO exceptions as of report date."] + [""] * (len(COLS_EXC) - 1), "grand")]
    it = items.sort_values(["Buyer", "DaysLate"], ascending=[True, False])
    rows = [([r[c[0]] for c in COLS_EXC], "detail") for _, r in it.iterrows()]
    tot = [""] * len(COLS_EXC)
    tot[0] = "GRAND TOTAL"
    tot[4] = f"{len(it)} line(s)"
    rows.append((tot, "grand"))
    return rows


def query_po_exceptions(include_leadtime=True, project_ids=None, date_from=None, date_to=None,
                        all_statuses=False):
    buyer_sel = "COALESCE(bu.EmpLastName + ', ' + bu.EmpFirstName, CAST(poh.BuyerID AS varchar(20)))"
    buyer_join = "LEFT JOIN tblEmployee bu ON bu.EmployeeID = poh.BuyerID"
    lead_sel = "eim.EstimatedLeadTime" if include_leadtime else "CAST(NULL AS int)"
    llt_sel = f"eim.[{_LLT_FLAG_COL}]" if include_leadtime else "CAST(NULL AS bit)"
    over_sel = f"eim.[{_OVERSIZE_FLAG_COL}]" if include_leadtime else "CAST(NULL AS bit)"
    rel_sel = f"CAST(eim.[{_ENG_RELEASE_COL}] AS date)" if include_leadtime else "CAST(NULL AS date)"
    lead_join = "LEFT JOIN tblEngItemMaster eim ON eim.ItemID = pod.ItemID" if include_leadtime else ""
    proj = ""
    if project_ids:
        ids = ",".join(str(int(p)) for p in project_ids)
        proj = f" AND pod.ProjectID IN ({ids})"
    return f"""
    SELECT
        {buyer_sel}                     AS Buyer,
        pod.ProjectID                   AS ProjectID,
        p.DisplayName                   AS JobName,
        pod.SpecID                      AS Code,
        pod.ItemID                      AS Item,
        pod.ItemDescription             AS Description,
        pdd.ItemMasterCategoryDescription AS Category,
        poh.PurchaseOrderID             AS PO,
        poh.CName                       AS Vendor,
        pod.PurchaseQty                 AS Qty,
        pod.Received                    AS Received,
        pod.ExtendedPrice               AS ExtValue,
        CAST(pod.DateRequired AS date)  AS DateRequired,
        CAST(pod.DateRevised  AS date)  AS DateRevised,
        CAST(pdd.LastReceivedDate AS date) AS ReceiptDate,
        CAST(poh.PurchaseDate AS date)  AS Ordered,
        {lead_sel}                      AS LeadDays,
        {llt_sel}                       AS LLTFlag,
        {over_sel}                      AS OverFlag,
        {rel_sel}                       AS EngReleaseDate
    FROM vwPurchaseOrderHeader poh
    JOIN vwPurchaseOrderDetails pod ON pod.PurchaseOrderID = poh.PurchaseOrderID
    LEFT JOIN vwPurchaseOrderDetailsDetailed pdd ON pdd.PurchaseDetailID = pod.PurchaseDetailID
    LEFT JOIN tblProjects p ON p.ProjectID = pod.ProjectID
    {buyer_join}
    {lead_join}
    WHERE poh.PurchaseActive = 1{"" if all_statuses else _RECV_OPEN_CLAUSE}{proj}{_po_date_window(date_from, date_to)}
    ORDER BY Buyer, pod.ProjectID, poh.PurchaseOrderID, pod.ItemID
    """


# open = not fully received; dropped when all_statuses=True (PurchaseActive=1 already excludes cancelled)
_RECV_OPEN_CLAUSE = "\n      AND (pod.Received IS NULL OR pod.Received < pod.PurchaseQty)"


def po_listing_detail(df, today=None):
    """One normalized row per PO LINE across ALL statuses (open, overdue, partial, received) on
    active — i.e. not cancelled — POs. Same 19-field layout and field sourcing as the exception
    detail; the difference is (a) no overdue filter and (b) a fuller Status: Received / Received —
    partial / Overdue / Overdue — partial / Open (before need-by). DaysLate is signed (negative =
    days until need-by) so the report can sort most-overdue first."""
    import pandas as pd
    today = today or _dt.date.today()
    out = []
    for _, r in df.iterrows():
        req = pd.to_datetime(r.get("DateRequired"), errors="coerce")
        rev = pd.to_datetime(r.get("DateRevised"), errors="coerce")
        need = rev if pd.notna(rev) else req
        needd = need.date() if pd.notna(need) else None
        rcpt = pd.to_datetime(r.get("ReceiptDate"), errors="coerce")
        eng = pd.to_datetime(r.get("EngReleaseDate"), errors="coerce")
        qty = _num_or_none(r.get("Qty")) or 0.0
        recv = _num_or_none(r.get("Received")) or 0.0
        full = qty > 0 and recv >= qty
        part = recv > 0 and not full
        overdue = bool(needd and needd < today and not full)
        if full:
            status = "Received"                                      # fully received / closed
        elif overdue:
            status = "Overdue — partial" if part else "Overdue"
        else:
            status = "Open — partial" if part else "Open"
        pid = r.get("ProjectID")
        out.append({
            "Buyer": (str(r.get("Buyer")).strip() if not _blank(r.get("Buyer")) else "(unassigned)"),
            "ProjectID": ("" if _blank(pid) else str(int(pid)) if _num_or_none(pid)
                          and float(pid).is_integer() else str(pid)),
            "JobName": ("" if _blank(r.get("JobName")) else str(r.get("JobName"))),
            "Code": _spec_code(r.get("Code")),
            "Item": ("" if _blank(r.get("Item")) else str(r.get("Item"))),
            "Category": ("" if _blank(r.get("Category")) else str(r.get("Category"))),
            "EngRelease": (eng.date().isoformat() if pd.notna(eng) else ""),
            "PO": ("" if _blank(r.get("PO")) else str(r.get("PO"))),
            "PlannedShip": "",                            # ETO holds no maintained ship date
            "PlannedReceipt": (req.date().isoformat() if pd.notna(req) else ""),
            "RevisedReceipt": (rev.date().isoformat() if pd.notna(rev) else ""),
            "ReceiptDate": (rcpt.date().isoformat() if pd.notna(rcpt) else ""),
            "Status": status,
            "LastUpdated": "",                            # no PO-line modified timestamp in ETO
            "DaysToAssembly": None,                       # no maintained assembly date
            "RFQDate": "",                                # not in ETO
            "PermitDates": "",                            # not in ETO
            "LeadTime": (int(_num_or_none(r.get("LeadDays"))) if _num_or_none(r.get("LeadDays")) else None),
            "Oversized": ("yes" if _flag(r.get("OverFlag")) else ""),
            "ExtValue": round(_num_or_none(r.get("ExtValue")) or 0.0, 2),
            "DaysLate": ((today - needd).days if needd else -10 ** 6),
        })
    return pd.DataFrame(out)


def po_listing_build_rows(items):
    """(cells, kind) rows for the all-status PO listing — sorted buyer, then most-overdue first."""
    if items is None or items.empty:
        return [(["No purchase-order lines for the selection."] + [""] * (len(COLS_EXC) - 1), "grand")]
    it = items.sort_values(["Buyer", "DaysLate"], ascending=[True, False])
    rows = [([r[c[0]] for c in COLS_EXC], "detail") for _, r in it.iterrows()]
    tot = [""] * len(COLS_EXC)
    tot[0] = "GRAND TOTAL"
    tot[4] = f"{len(it)} line(s)"
    rows.append((tot, "grand"))
    return rows


def _blank(v):
    import pandas as pd
    if v is None:
        return True
    if isinstance(v, float) and pd.isna(v):
        return True
    return str(v).strip().lower() in ("", "nan", "none")


def _flag(v):
    import pandas as pd
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return False
    if isinstance(v, str):
        return v.strip().lower() in ("1", "true", "y", "yes", "t")
    return bool(v)


def exc_classify(df, today=None):
    import pandas as pd
    today = today or _dt.date.today()
    out = []
    for _, r in df.iterrows():
        req = pd.to_datetime(r.get("DateRequired"), errors="coerce")
        rev = pd.to_datetime(r.get("DateRevised"), errors="coerce")
        need = (rev if pd.notna(rev) else req)
        need = need.date() if pd.notna(need) else None
        eng_rel = pd.to_datetime(r.get("EngReleaseDate"), errors="coerce")
        eng_rel = eng_rel.date() if pd.notna(eng_rel) else None
        # Delivered-late (overdue) is the only ETO-consistent lateness signal on open lines.
        # Ord Late / Critical were derived from EstimatedLeadTime, which is unmaintained
        # (always NULL) — dropped. LLT / Oversize stay: they're maintained ETO flags.
        del_late = bool(need and need < today)
        days_late = (today - need).days if del_late else 0
        llt = _flag(r.get("LLTFlag"))
        oversized = _flag(r.get("OverFlag"))
        if not del_late:
            continue
        rec = dict(r)
        rec.update(Buyer=(str(r.get("Buyer")).strip() if not _blank(r.get("Buyer")) else "(unassigned)"),
                   NeedBy=need, EngRel=eng_rel, LLT=llt, Oversized=oversized,
                   DelLate=del_late, DaysLate=days_late)
        out.append(rec)
    return pd.DataFrame(out)


def exc_aggregate(ex):
    import pandas as pd
    if ex.empty:
        return ex

    def _join(series):
        seen, out = set(), []
        for v in series:
            s = str(v).strip()
            if s and s.lower() not in ("nan", "none") and s not in seen:
                seen.add(s); out.append(s)
        return ", ".join(out)

    rows = []
    for (pid, item), g in ex.groupby(["ProjectID", "Item"], dropna=False):
        need_vals = [d for d in g["NeedBy"] if d]
        rows.append({
            "Buyer": _join(g["Buyer"]),
            "ProjectID": "" if pd.isna(pid) else str(int(pid)) if float(pid).is_integer() else str(pid),
            "JobName": next((j for j in g["JobName"] if not _blank(j)), ""),
            "Item": "" if _blank(item) else str(item),
            "Description": next((d for d in g["Description"] if not _blank(d)), ""),
            "POs": _join(g["PO"]),
            "Vendors": _join(g["Vendor"]),
            "POCount": int(g["PO"].nunique()),
            "ExtValue": round(float(g["ExtValue"].fillna(0).sum()), 2),
            "NeedBy": min(need_vals).isoformat() if need_vals else "",
            "DaysLate": int(g["DaysLate"].max()),
            "EngRelease": next((d.isoformat() for d in g["EngRel"] if d), ""),
            "LLT": "LLT" if g["LLT"].any() else "",
            "Oversized": "OVER" if g["Oversized"].any() else "",
            "DelLate": "LATE" if g["DelLate"].any() else "",
        })
    out = pd.DataFrame(rows)
    return out.sort_values(["Buyer", "DaysLate"], ascending=[True, False]).reset_index(drop=True)


def exc_build_rows(items):
    if items.empty:
        return [(["No open-PO exceptions as of report date."] + [""] * (len(COLS_FLAT) - 1), "grand")]
    rows = [([r[c[0]] for c in COLS_FLAT], "detail") for _, r in items.iterrows()]
    rows.append((["GRAND TOTAL", "", "", f"{len(items)} item(s)", "", "", "",
                  int(items["POCount"].sum()), round(float(items["ExtValue"].sum()), 2)]
                 + [""] * (len(COLS_FLAT) - 9), "grand"))
    return rows


def exceptions_book_bytes(items, label):
    """Single AutoFiltered sheet. items = exc_detail output (one row per open, overdue PO line)."""
    from openpyxl.worksheet.properties import PageSetupProperties
    rows = exc_detail_build_rows(items)
    wb = Workbook(); wb.remove(wb.active)
    ws = wb.create_sheet("Procurement Exceptions")
    _populate_sheet(ws, "Procurement Exceptions by Buyer",
                    "Open POs past need-by — one row per PO line (sortable)",
                    COLS_EXC, rows, label)
    last_data = 3 + (len(items) if items is not None and not items.empty else 0)
    if last_data >= 4:
        ws.auto_filter.ref = f"A3:{get_column_letter(len(COLS_EXC))}{last_data}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# LATE VENDORS — OVERDUE open PO lines, by vendor (needs no receiver log).
# Applies ETO's lateness definition (from dbo.urpPurchasingLateVendors:
# need-by = ISNULL(DateRevised, DateRequired)) to the STILL-OPEN population:
#   open      = pod.Received < pod.PurchaseQty
#   overdue   = need-by < today
#   DaysLate  = today − need-by
# Background: ETO's native urpPurchasingLateVendors measures the RECEIVED population
# instead — DaysLate = receipt date − need-by, where the receipt date comes from
# vwReceiverLogSummed.MaxOfDate. That receipt date is the ONLY thing the receiver log
# provides here (pod has a received qty, no date), so the historical "arrived late"
# case stays with ETO's own report; this console report covers what's overdue NOW.
# ─────────────────────────────────────────────────────────────────────────────
COLS_LATE = [
    ("ProjectID",   "Project",     8,  "C", False),
    ("PO",          "PO #",        7,  "L", False),
    ("Item",        "Item",        10, "L", False),
    ("Description", "Description", 26, "L", False),
    ("Qty",         "Qty",          6, "R", True),
    ("Received",    "Rec'd",        6, "R", True),
    ("Required",    "Required",    10, "C", False),
    ("Revised",     "Revised",     10, "C", False),
    ("DaysLate",    "Days Late",    8, "R", True),
    ("ExtValue",    "Ext. Value",  11, "R", True),
]


def query_late_vendors(project_ids=None, date_from=None, date_to=None):
    """OVERDUE open PO lines, by vendor. A line still open (Received < PurchaseQty) whose
    need-by (ISNULL(DateRevised, DateRequired)) is already past is late NOW — Days Late =
    today − need-by. This applies urpPurchasingLateVendors' lateness definition to the
    not-yet-received population, so no receipt date / receiver log is needed. (The historical
    'received but arrived late' case needs the receiver log's receipt date — ETO's own report.)"""
    proj = ""
    if project_ids:
        ids = ",".join(str(int(p)) for p in project_ids)
        proj = f" AND pod.ProjectID IN ({ids})"
    need_by = "ISNULL(pod.DateRevised, pod.DateRequired)"
    dt = _po_date_window(date_from, date_to)
    return f"""
    SELECT
        poh.CName                        AS Supplier,
        pod.ProjectID                    AS ProjectID,
        poh.PurchaseOrderID              AS PO,
        COALESCE(bu.EmpLastName + ', ' + bu.EmpFirstName,
                 CAST(poh.BuyerID AS varchar(20)))       AS Buyer,
        pod.ItemID                       AS Item,
        pod.ItemDescription              AS Description,
        pod.PurchaseQty                  AS Qty,
        pod.Received                     AS Received,
        CAST(pod.DateRequired AS date)   AS Required,
        CAST(pod.DateRevised  AS date)   AS Revised,
        DATEDIFF(d, {need_by}, CAST(GETDATE() AS date)) AS DaysLate,
        pod.ExtendedPrice                AS ExtValue,
        p.DisplayName                    AS JobName,
        p.PStatus                        AS ProjStatus
    FROM vwPurchaseOrderHeader poh
    JOIN vwPurchaseOrderDetails pod ON pod.PurchaseOrderID = poh.PurchaseOrderID
    LEFT JOIN tblProjects p ON p.ProjectID = pod.ProjectID
    LEFT JOIN tblEmployee bu ON bu.EmployeeID = poh.BuyerID
    WHERE poh.PurchaseActive = 1
      AND (pod.Received IS NULL OR pod.Received < pod.PurchaseQty)
      AND {need_by} < CAST(GETDATE() AS date){proj}{dt}
    ORDER BY poh.CName, DaysLate DESC
    """


def query_po_totals_by_project(project_ids=None, date_from=None, date_to=None):
    """Per-project purchase totals over ALL active PO lines — closed (received) AND open —
    scoped by the PO-placed window. Total = received + open; overdue is the open-past-need-by
    subset. Values in CAD (rate≤0 guarded to 1)."""
    proj = ""
    if project_ids:
        ids = ",".join(str(int(p)) for p in project_ids)
        proj = f" AND pod.ProjectID IN ({ids})"
    dt = _po_date_window(date_from, date_to)
    rate = "CASE WHEN poh.PurchaseCurrRate > 0 THEN poh.PurchaseCurrRate ELSE 1 END"
    ext = f"pod.ExtendedPrice * {rate}"
    open_ = "(pod.Received IS NULL OR pod.Received < pod.PurchaseQty)"
    closed = "(pod.Received IS NOT NULL AND pod.Received >= pod.PurchaseQty)"
    needby = "ISNULL(pod.DateRevised, pod.DateRequired)"
    overdue = f"({open_} AND {needby} < CAST(GETDATE() AS date))"
    return f"""
    SELECT
        pod.ProjectID                                       AS ProjectID,
        MAX(p.DisplayName)                                  AS JobName,
        COUNT(DISTINCT poh.PurchaseOrderID)                 AS POs,
        COUNT(*)                                            AS Lines,
        SUM(CASE WHEN {open_}   THEN 1 ELSE 0 END)          AS OpenLines,
        SUM(CASE WHEN {closed}  THEN 1 ELSE 0 END)          AS ClosedLines,
        CAST(SUM({ext}) AS decimal(20,2))                                        AS TotalPurchases,
        CAST(SUM(CASE WHEN {closed}  THEN {ext} ELSE 0 END) AS decimal(20,2))    AS ReceivedValue,
        CAST(SUM(CASE WHEN {open_}   THEN {ext} ELSE 0 END) AS decimal(20,2))    AS OpenValue,
        CAST(SUM(CASE WHEN {overdue} THEN {ext} ELSE 0 END) AS decimal(20,2))    AS OverdueValue
    FROM vwPurchaseOrderHeader poh
    JOIN vwPurchaseOrderDetails pod ON pod.PurchaseOrderID = poh.PurchaseOrderID
    LEFT JOIN tblProjects p ON p.ProjectID = pod.ProjectID
    WHERE poh.PurchaseActive = 1{proj}{dt}
    GROUP BY pod.ProjectID
    ORDER BY TotalPurchases DESC
    """


def query_po_by_buyer(project_ids=None, date_from=None, date_to=None):
    """Per-buyer purchasing rollup over active PO lines (scoped by project + PO-placed window):
    POs, lines, committed value (CAD), and the open / overdue subset. The buyer-centric view."""
    proj = ""
    if project_ids:
        ids = ",".join(str(int(p)) for p in project_ids)
        proj = f" AND pod.ProjectID IN ({ids})"
    dt = _po_date_window(date_from, date_to)
    buyer = ("COALESCE(bu.EmpLastName + ', ' + bu.EmpFirstName, "
             "CAST(poh.BuyerID AS varchar(20)))")
    rate = "CASE WHEN poh.PurchaseCurrRate > 0 THEN poh.PurchaseCurrRate ELSE 1 END"
    open_ = "(pod.Received IS NULL OR pod.Received < pod.PurchaseQty)"
    needby = "ISNULL(pod.DateRevised, pod.DateRequired)"
    overdue = f"({open_} AND {needby} < CAST(GETDATE() AS date))"
    return f"""
    SELECT
        {buyer}                                             AS Buyer,
        COUNT(DISTINCT poh.PurchaseOrderID)                 AS POs,
        COUNT(*)                                            AS Lines,
        CAST(SUM(pod.ExtendedPrice * {rate}) AS decimal(20,2)) AS ExtValue,
        SUM(CASE WHEN {open_} THEN 1 ELSE 0 END)            AS OpenLines,
        SUM(CASE WHEN {overdue} THEN 1 ELSE 0 END)          AS OverdueLines,
        CAST(SUM(CASE WHEN {overdue} THEN pod.ExtendedPrice * {rate} ELSE 0 END)
             AS decimal(20,2))                              AS OverdueValue
    FROM vwPurchaseOrderHeader poh
    JOIN vwPurchaseOrderDetails pod ON pod.PurchaseOrderID = poh.PurchaseOrderID
    LEFT JOIN tblEmployee bu ON bu.EmployeeID = poh.BuyerID
    WHERE poh.PurchaseActive = 1{proj}{dt}
    GROUP BY {buyer}
    ORDER BY OverdueValue DESC, ExtValue DESC
    """


def _ds(v):
    """NaN/None/NaT-safe date string (blank when missing)."""
    if v is None:
        return ""
    try:
        if v != v:            # NaN
            return ""
    except Exception:
        pass
    try:
        return v.strftime("%Y-%m-%d")
    except Exception:
        s = str(v)
        return "" if s.strip().lower() in ("nan", "nat", "none") else s


def late_build_rows(df):
    """Group by vendor: band → overdue lines (worst first) → per-vendor subtotal → grand."""
    n = len(COLS_LATE)
    if df.empty:
        return [(["No overdue open PO lines for the selection."] + [""] * (n - 1), "grand")]
    rows = []
    for sup in sorted(df["Supplier"].dropna().unique(), key=str):
        ssub = df[df["Supplier"] == sup].sort_values("DaysLate", ascending=False)
        rows.append(([f"Vendor: {sup}   ·   {len(ssub)} overdue line(s)"] + [""] * (n - 1), "l3_sub"))
        for _, r in ssub.iterrows():
            rows.append(([r.ProjectID, r.PO, r.Item, r.Description, r.Qty, r.Received,
                          _ds(r.Required), _ds(r.Revised),
                          int(r.DaysLate), round(float(r.ExtValue), 2)], "detail"))
        rows.append(([f"{sup} — Overdue Value", "", "", "", "", "", "", "",
                      int(ssub["DaysLate"].max()), round(float(ssub["ExtValue"].sum()), 2)], "l1_sub"))
    rows.append((["GRAND TOTAL — Overdue Value"] + [""] * 7
                 + [int(df["DaysLate"].max()), round(float(df["ExtValue"].sum()), 2)], "grand"))
    return rows


def late_vendors_book_bytes(df, label):
    rows = late_build_rows(df)
    wb = Workbook(); wb.remove(wb.active)
    ws = wb.create_sheet("Late Vendors")
    _populate_sheet(ws, "Purchasing — Late Vendors",
                    "Open PO lines past their need-by (revised, else required) — overdue, by vendor",
                    COLS_LATE, rows, label)
    ci = [i for i, c in enumerate(COLS_LATE, 1) if c[0] == "ExtValue"][0]
    for row in range(4, ws.max_row + 1):
        ws.cell(row=row, column=ci).number_format = "$#,##0.00"
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# LATE VENDORS (DELIVERED LATE) — the urpPurchasingLateVendors logic, our parameters.
# RECEIVED lines that arrived after need-by, by vendor (the vendor delivery scorecard).
#   need-by  = ISNULL(DateRevised, DateRequired)
#   receipt  = vwReceiverLogSummed.MaxOfDate   (actual last receipt date — the SP's source)
#   DaysLate = receipt − need-by   (> 0);  fully received: PurchaseQty − SumOfQtyReceived <= 0
# Built on the PO views + vwReceiverLogSummed so we can scope by the selected projects and by
# a PO-created window (@datPOCreated*) — the parameters the proc itself doesn't take. If the
# account lacks direct SELECT on vwReceiverLogSummed, the service falls back to EXEC_LATE_VENDORS.
# ─────────────────────────────────────────────────────────────────────────────
COLS_DELIVERED = [
    ("ProjectID",   "Project",     8,  "C", False),
    ("PO",          "PO #",        7,  "L", False),
    ("Item",        "Item",        12, "L", False),
    ("Description", "Description", 24, "L", False),
    ("Qty",         "Qty",          6, "R", True),
    ("QtyReceived", "Rec'd",        6, "R", True),
    ("Required",    "Required",    10, "C", False),
    ("Revised",     "Revised",     10, "C", False),
    ("Received",    "Received",    10, "C", False),
    ("DaysLate",    "Days Late",    8, "R", True),
    ("ExtValue",    "Ext. Value",  11, "R", True),
]

# EXEC fallback: the proc bounds by PO-created window; projects are filtered client-side.
EXEC_LATE_VENDORS = ("SET NOCOUNT ON; EXEC dbo.urpPurchasingLateVendors "
                     "@nvcCompanyIDIn=NULL, @datPOCreatedLower=?, @datPOCreatedUpper=?")
# proc output column → our delivered-late column
EXEC_COLMAP = {"SupplierName": "Supplier", "ProjectID": "ProjectID", "PurchaseOrderID": "PO",
               "ItemCompanyID": "Item", "ItemDescription": "Description", "PurchaseQty": "Qty",
               "QtyReceived": "QtyReceived", "DateRequired": "Required", "DateRevised": "Revised",
               "MaxOfDate": "Received", "DaysLate": "DaysLate", "ExtendedPriceExchange": "ExtValue"}


def query_delivered_late(project_ids=None, date_from=None, date_to=None):
    """Our own project/date-scoped transcription of urpPurchasingLateVendors."""
    proj = ""
    if project_ids:
        ids = ",".join(str(int(p)) for p in project_ids)
        proj = f" AND pod.ProjectID IN ({ids})"
    win = ""
    if date_from:
        win += f" AND poh.PurchaseDate >= '{date_from}'"
    if date_to:
        win += f" AND poh.PurchaseDate < DATEADD(day, 1, '{date_to}')"
    need_by = "ISNULL(pod.DateRevised, pod.DateRequired)"
    return f"""
    SELECT
        poh.CName                        AS Supplier,
        pod.ProjectID                    AS ProjectID,
        poh.PurchaseOrderID              AS PO,
        ISNULL(pod.ItemCompanyID, CAST(pod.ItemID AS nvarchar(30))) AS Item,
        pod.ItemDescription              AS Description,
        pod.PurchaseQty                  AS Qty,
        CAST(rls.SumOfQtyReceived AS decimal(20,6)) AS QtyReceived,
        CAST(pod.DateRequired AS date)   AS Required,
        CAST(pod.DateRevised  AS date)   AS Revised,
        CAST(rls.MaxOfDate    AS date)   AS Received,
        DATEDIFF(d, {need_by}, rls.MaxOfDate) AS DaysLate,
        CAST(pod.PurchaseQty * pod.PurchasePrice * poh.PurchaseCurrRate AS decimal(20,6)) AS ExtValue,
        p.DisplayName                    AS JobName,
        p.PStatus                        AS ProjStatus
    FROM vwPurchaseOrderHeader poh
    JOIN vwPurchaseOrderDetails pod ON pod.PurchaseOrderID = poh.PurchaseOrderID
    JOIN vwReceiverLogSummed rls ON rls.PurchaseDetailID = pod.PurchaseDetailID
    LEFT JOIN tblProjects p ON p.ProjectID = pod.ProjectID
    WHERE poh.PurchaseActive = 1
      AND (pod.PurchaseQty - ISNULL(rls.SumOfQtyReceived, 0)) <= 0
      AND {need_by} IS NOT NULL
      AND DATEDIFF(d, {need_by}, rls.MaxOfDate) > 0{proj}{win}
    ORDER BY poh.CName, DaysLate DESC
    """


def delivered_build_rows(df):
    """Group by vendor: band → late lines (worst first) → per-vendor subtotal → grand."""
    n = len(COLS_DELIVERED)
    if df.empty:
        return [(["No delivered-late lines in the selected window."] + [""] * (n - 1), "grand")]
    rows = []
    for sup in sorted(df["Supplier"].dropna().unique(), key=str):
        ssub = df[df["Supplier"] == sup].sort_values("DaysLate", ascending=False)
        avg = ssub["DaysLate"].mean()
        rows.append(([f"Vendor: {sup}   ·   {len(ssub)} late line(s) · avg {avg:.0f}d"] + [""] * (n - 1), "l3_sub"))
        for _, r in ssub.iterrows():
            rows.append(([r.ProjectID, r.PO, r.Item, r.Description, r.Qty, r.QtyReceived,
                          _ds(r.Required), _ds(r.Revised), _ds(r.Received),
                          int(r.DaysLate), round(float(r.ExtValue), 2)], "detail"))
        rows.append(([f"{sup} — Late Value", "", "", "", "", "", "", "", "",
                      int(ssub["DaysLate"].max()), round(float(ssub["ExtValue"].sum()), 2)], "l1_sub"))
    rows.append((["GRAND TOTAL — Late Value"] + [""] * 8
                 + [int(df["DaysLate"].max()), round(float(df["ExtValue"].sum()), 2)], "grand"))
    return rows


def delivered_book_bytes(df, label):
    rows = delivered_build_rows(df)
    wb = Workbook(); wb.remove(wb.active)
    ws = wb.create_sheet("Late Vendors")
    _populate_sheet(ws, "Purchasing — Late Vendors (Delivered Late)",
                    "Received lines that arrived after need-by (revised, else required) — by vendor",
                    COLS_DELIVERED, rows, label)
    ci = [i for i, c in enumerate(COLS_DELIVERED, 1) if c[0] == "ExtValue"][0]
    for row in range(4, ws.max_row + 1):
        ws.cell(row=row, column=ci).number_format = "$#,##0.00"
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# Web adapters — turn (col_defs, grouped rows) into the console's generic shape
# ─────────────────────────────────────────────────────────────────────────────
_MONEY_FIELDS = {"LabourCost", "Cost", "WagesPayable", "ExtValue", "Price"}
_INT_FIELDS = {"Entries", "UniqueEmps", "Employees", "POCount", "DaysLate", "Lead",
               "DaysToAssembly", "LeadTime"}
_WRAP_FIELDS = {"JobName", "Description"}


def web_type(field, is_numeric):
    if not is_numeric:
        return "text"
    if field in _MONEY_FIELDS:
        return "money2"          # currency WITH cents (matches the workbook)
    if field in _INT_FIELDS:
        return "int"
    return "num"


def web_columns(col_defs):
    """col_defs -> list of (key,label,type,align,wrap) tuples for QueryColumn."""
    out = []
    for field, header, _w, align, is_num in col_defs:
        out.append((field, header, web_type(field, is_num),
                    "right" if align == "R" else "left", field in _WRAP_FIELDS))
    return out


def _native(v):
    """Coerce numpy/pandas scalars to plain Python so the row dict is JSON-serializable."""
    if v is None or v == "":
        return None
    item = getattr(v, "item", None)   # numpy scalar -> python scalar
    if callable(item):
        try:
            return v.item()
        except Exception:
            pass
    return v


def web_rows(col_defs, grouped_rows):
    """grouped (cells, kind) tuples -> list of row dicts carrying a reserved _kind."""
    keys = [c[0] for c in col_defs]
    out = []
    for cells, kind in grouped_rows:
        d = {k: _native(v) for k, v in zip(keys, cells)}
        d["_kind"] = kind
        out.append(d)
    return out
