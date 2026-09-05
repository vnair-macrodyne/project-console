"""
Exporters — turn any QueryResult into a downloadable Excel workbook or PDF.

Both are generic: they read the QueryResult's typed columns and format each cell
by column type (hours / money / pct / date / int / text), so a new query needs no
exporter changes. Output is tenant-branded (product + company + header colour).
"""
import io
from datetime import datetime

from console_web.queries import branding
from console_web import etospec


# ─────────────────────────────────────────────────────────────────────────────
# value formatting shared by both exporters
# ─────────────────────────────────────────────────────────────────────────────
def _fmt(value, ctype):
    if value is None or value == "":
        return ""
    try:
        if ctype == "hours":
            return "{:,.0f}".format(float(value))
        if ctype == "num":
            f = float(value)
            return "{:,.0f}".format(f) if f == int(f) else "{:,.2f}".format(f)
        if ctype == "money" or ctype == "money2":
            return "${:,.2f}".format(float(value))
        if ctype == "pct":
            return "{:.1f}%".format(float(value) * 100)
        if ctype == "int":
            return "{:,d}".format(int(float(value)))
        if ctype == "id":
            return str(int(float(value)))
    except (TypeError, ValueError):
        return str(value)
    return str(value)


def _ts():
    # container clock only; label, not logic
    return datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")


# Executive-dashboard palette (mirrors console_dashboard.py)
C_GROUP = "2E75B6"   # medium-blue block band
C_WARN = "FFEB9C"    # 90–100% amber
C_BAD = "FFC7CE"     # >100% / slipped red
C_ZEBRA = "F2F4F8"


def _tl_hex(ctype, value):
    """Traffic-light fill (amber/red) for a cell — matches the workbook rules."""
    if value is None:
        return None
    try:
        v = float(value)
    except (TypeError, ValueError):
        return None
    if ctype == "pct":
        if v > 1.0:
            return C_BAD
        if v >= 0.9:
            return C_WARN
    elif ctype == "days":
        if v > 0:
            return C_BAD
    return None


def _blocks(columns):
    """Contiguous (block_label, start_index, span) runs across the columns."""
    runs = []
    i = 0
    while i < len(columns):
        blk = columns[i].block
        span = 1
        while i + span < len(columns) and columns[i + span].block == blk:
            span += 1
        runs.append((blk, i, span))
        i += span
    return runs


# ─────────────────────────────────────────────────────────────────────────────
# Excel
# ─────────────────────────────────────────────────────────────────────────────
def _spec_xlsx(exp) -> bytes:
    """Faithful workbook for a deployed eto-reporting report (labour / PO), built by
    the vendored writer so the download matches the on-prem report exactly."""
    kind = exp["kind"]
    if kind == "labour":
        return etospec.labour_book_bytes(exp["report_id"], exp["rows"], exp["label"])
    if kind == "po_status":
        return etospec.po_status_book_bytes(exp["df"], exp["label"])
    if kind == "exceptions":
        return etospec.exceptions_book_bytes(exp["items"], exp["label"])
    if kind == "late":
        return etospec.late_vendors_book_bytes(exp["df"], exp["label"])
    if kind == "delivered":
        return etospec.delivered_book_bytes(exp["df"], exp["label"])
    raise ValueError(f"unknown spec export kind '{kind}'")


class _Col:
    """Lightweight column descriptor for the flattened export (decoupled from QueryColumn)."""
    __slots__ = ("key", "label", "type", "align", "calc")

    def __init__(self, key, label, ctype="text", align="left", calc=False):
        self.key, self.label, self.type, self.align, self.calc = key, label, ctype, align, calc


# Row kinds: bands we CARRY DOWN into leading columns, and subtotal/total rows we DROP.
_CARRY_BANDS = ("l3_sub", "l2_sub")          # outer → inner grouping levels
_DROP_KINDS = ("l1_sub", "grand", "l2_sub", "l3_sub")

# Band-label prefixes we recognise, so a leading column gets a real name and a clean value
# (e.g. "Machine 12" → header "Machine", value "12"). Job names etc. never match, so they're
# left whole under a generic header rather than being mis-split.
_KNOWN_PREFIXES = ("Project", "Machine", "Vendor", "Supplier", "Buyer", "Discipline", "Category")


def _band_label(row, cols):
    """The visible label carried by a band row = first non-empty cell (skipping helper keys)."""
    for c in cols:
        v = row.get(c.key)
        if v not in (None, ""):
            return v.strip() if isinstance(v, str) else str(v)
    for k, v in row.items():
        if not k.startswith("_") and v not in (None, ""):
            return str(v)
    return ""


def _derive_group_header(label, default):
    """('Project: 240088' → 'Project', ':'), ('Machine 12' → 'Machine', ' '), else (default, None)."""
    if ":" in label:
        pre = label.split(":", 1)[0].strip()
        if pre:
            return pre, ":"
    first = label.split(" ", 1)[0].strip()
    if first in _KNOWN_PREFIXES and " " in label:
        return first, " "
    return default, None


def _flatten(result):
    """Turn a (possibly banded / block-grouped) QueryResult into a flat, pivot-friendly
    (columns, rows) pair: one row per detail line, every column populated, grouping context
    carried into leading columns, and subtotal/total rows removed. No column merges, no
    subheads — so users can sort, subtotal and pivot freely."""
    cols = result.columns
    rows = result.rows or []

    present = [lv for lv in _CARRY_BANDS if any(r.get("_kind") == lv for r in rows)]
    # Decide each carried level's leading-column header once, from its first band label.
    meta = {}   # level -> (synthetic_key, header, separator)
    group_cols = []
    for i, lv in enumerate(present):
        first = next((_band_label(r, cols) for r in rows if r.get("_kind") == lv), "")
        hdr, sep = _derive_group_header(first, "Group" if i == 0 else f"Group {i + 1}")
        key = f"_grp{i}"
        meta[lv] = (key, hdr, sep)
        group_cols.append(_Col(key, hdr, "text", "left"))

    ctx = {lv: "" for lv in present}
    flat_rows = []
    for r in rows:
        kind = r.get("_kind", "detail")
        if kind in present:
            key, hdr, sep = meta[kind]
            lbl = _band_label(r, cols)
            if sep and sep in lbl:
                lbl = lbl.split(sep, 1)[1].strip()
            ctx[kind] = lbl
            for deeper in present[present.index(kind) + 1:]:   # reset inner levels
                ctx[deeper] = ""
            continue
        if kind in _DROP_KINDS:                                 # subtotal / grand — dropped
            continue
        nr = {meta[lv][0]: ctx[lv] for lv in present}
        nr.update({c.key: r.get(c.key) for c in cols})
        flat_rows.append(nr)

    # Detail columns keep their key/type; block name (if any) is folded into the label so a
    # single header row stays unambiguous ("Labour — Budget"). A grouping column that was
    # hoisted into a leading column (and is now blank on every detail row) is dropped as noise.
    group_headers = {meta[lv][1] for lv in present}
    detail_cols = []
    for c in cols:
        if c.label in group_headers and all(r.get(c.key) in (None, "") for r in flat_rows):
            continue
        detail_cols.append(_Col(c.key,
                                (f"{c.block} — {c.label}" if getattr(c, "block", "") else c.label),
                                c.type, c.align, getattr(c, "calc", False)))
    return group_cols + detail_cols, flat_rows


def _put_cell(cell, raw, ctype):
    """Write a value with the right number format so numbers stay numeric (pivot-friendly)."""
    if raw is None or raw == "":
        cell.value = ""
        return
    if ctype in ("hours", "int", "days") and isinstance(raw, (int, float)):
        cell.value = raw
        cell.number_format = "#,##0"
    elif ctype == "num" and isinstance(raw, (int, float)):
        cell.value = raw
        cell.number_format = "#,##0.##"
    elif ctype == "id" and isinstance(raw, (int, float)):
        cell.value = int(raw)
        cell.number_format = "0"
    elif ctype in ("money", "money2") and isinstance(raw, (int, float)):
        cell.value = raw
        cell.number_format = '"$"#,##0.00'
    elif ctype == "pct" and isinstance(raw, (int, float)):
        cell.value = raw
        cell.number_format = "0.0%"
    else:
        cell.value = raw if ctype in ("text", "date") else _fmt(raw, ctype)


def to_xlsx(result) -> bytes:
    """Flat, pivot-friendly workbook. Sheet 'Data' is a single clean table (headers in row 1,
    one row per line, grouping carried into leading columns, no subheads or subtotals, with an
    AutoFilter); sheet 'Info' carries the title, source and summary cards. This intentionally
    does NOT mimic the on-screen banded layout — it's built for sorting, subtotalling and pivots.
    (The faithful on-prem workbook writers remain in _spec_xlsx / etospec if ever needed.)"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    b = branding()
    hexcolor = (b["color"] or "1F3864").lstrip("#")
    cols, rows = _flatten(result)

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    head_fill = PatternFill("solid", fgColor=hexcolor)
    head_font = Font(name="Helvetica", bold=True, color="FFFFFF", size=10)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(bottom=thin)

    for j, col in enumerate(cols, start=1):
        c = ws.cell(1, j, col.label)
        c.fill = head_fill
        c.font = head_font
        c.alignment = Alignment(horizontal=col.align, vertical="center")
        c.border = border

    for i, row in enumerate(rows, start=2):
        for j, col in enumerate(cols, start=1):
            cell = ws.cell(i, j)
            _put_cell(cell, row.get(col.key), col.type)
            cell.alignment = Alignment(horizontal=col.align)

    for j, col in enumerate(cols, start=1):
        w = {"text": 20, "date": 12, "pct": 10, "money": 14, "money2": 14,
             "int": 10, "days": 10, "hours": 11, "num": 11, "id": 10}.get(col.type, 12)
        w = max(w, min(len(col.label) + 3, 34))
        ws.column_dimensions[get_column_letter(j)].width = w

    last_col = get_column_letter(max(len(cols), 1))
    last_row = max(len(rows) + 1, 1)
    ws.auto_filter.ref = f"A1:{last_col}{last_row}"
    ws.freeze_panes = "A2"

    # ── Info sheet: title, source, generated stamp, notes, summary cards ────────
    info = wb.create_sheet("Info")
    title_font = Font(name="Helvetica", bold=True, size=14, color=hexcolor)
    sub_font = Font(name="Helvetica", size=10, color="808080")
    lbl_font = Font(name="Helvetica", bold=True, size=10)
    info.cell(1, 1, f"{b['product']} — {result.title}").font = title_font
    info.cell(2, 1, f"{b['company']}   ·   generated {_ts()}").font = sub_font
    r = 4
    if result.cards:
        for card in result.cards:
            info.cell(r, 1, card.label).font = lbl_font
            info.cell(r, 2, str(card.value))
            r += 1
        r += 1
    if result.note:
        c = info.cell(r, 1, result.note)
        c.font = sub_font
        c.alignment = Alignment(wrap_text=True, vertical="top")
        info.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    info.column_dimensions["A"].width = 22
    info.column_dimensions["B"].width = 40

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# PDF
# ─────────────────────────────────────────────────────────────────────────────
def to_pdf(result) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib.units import inch
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                     Paragraph, Spacer)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    b = branding()
    hexcolor = "#" + (b["color"] or "1F3864").lstrip("#")
    brand = colors.HexColor(hexcolor)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(letter),
                            leftMargin=0.5 * inch, rightMargin=0.5 * inch,
                            topMargin=0.55 * inch, bottomMargin=0.5 * inch,
                            title=f"{b['product']} — {result.title}")
    styles = getSampleStyleSheet()
    h = ParagraphStyle("h", parent=styles["Title"], textColor=brand,
                       fontSize=16, spaceAfter=2, alignment=0)
    sub = ParagraphStyle("sub", parent=styles["Normal"], fontSize=8,
                         textColor=colors.grey, spaceAfter=2)
    cellst = ParagraphStyle("cell", parent=styles["Normal"], fontSize=7.5, leading=9)

    story = [Paragraph(f"{b['product']} — {result.title}", h),
             Paragraph(f"{b['company']} &nbsp;·&nbsp; generated {_ts()}", sub)]
    if result.cards:
        story.append(Paragraph(
            " &nbsp;&nbsp;|&nbsp;&nbsp; ".join(f"<b>{c.label}:</b> {c.value}"
                                               for c in result.cards), sub))
    story.append(Spacer(1, 8))

    cols = result.columns
    grouped = any(c.block for c in cols)
    if grouped:
        cellst = ParagraphStyle("cellg", parent=styles["Normal"], fontSize=6, leading=7.5)
    hcell = ParagraphStyle("hcell", parent=cellst, textColor=colors.white, fontSize=cellst.fontSize)

    data = []
    header_idx = 0
    if grouped:
        band = [""] * len(cols)
        for blk, start, span in _blocks(cols):
            if blk:
                band[start] = Paragraph(f"<b>{blk.upper()}</b>", hcell)
        data.append(band)
        header_idx = 1
    data.append([Paragraph(f"<b>{c.label}</b>", hcell) for c in cols])
    for row in result.rows:
        data.append([Paragraph(_fmt(row.get(col.key), col.type), cellst) for col in cols])
    data_start = header_idx + 1

    # column widths
    avail = landscape(letter)[0] - inch
    if grouped:
        wmap = {"text": 3.2, "date": 1.5, "pct": 1.3, "money": 1.7, "money2": 1.7, "int": 1.0, "days": 1.1}
        weights = [wmap.get(c.type, 1.2) for c in cols]
    else:
        weights = [max(len(c.label), 6) for c in cols]
    tot = sum(weights) or 1
    widths = [avail * w / tot for w in weights]

    table = Table(data, colWidths=widths, repeatRows=data_start)
    style = [
        ("BACKGROUND", (0, header_idx), (-1, header_idx), brand),
        ("TEXTCOLOR", (0, header_idx), (-1, header_idx), colors.white),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, header_idx), (-1, -1), 0.25, colors.HexColor("#E0E0E0")),
        ("TOPPADDING", (0, 0), (-1, -1), 2 if grouped else 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2 if grouped else 3),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
    ]
    if grouped:
        for blk, start, span in _blocks(cols):
            if blk:
                style.append(("SPAN", (start, 0), (start + span - 1, 0)))
                style.append(("BACKGROUND", (start, 0), (start + span - 1, 0),
                              colors.HexColor("#" + C_GROUP)))
        # zebra + per-cell traffic lights on the data region
        style.append(("ROWBACKGROUNDS", (0, data_start), (-1, -1),
                      [colors.white, colors.HexColor("#" + C_ZEBRA)]))
        for ri, row in enumerate(result.rows):
            for j, col in enumerate(cols):
                tl = _tl_hex(col.type, row.get(col.key))
                if tl:
                    style.append(("BACKGROUND", (j, data_start + ri),
                                  (j, data_start + ri), colors.HexColor("#" + tl)))
    else:
        style.append(("ROWBACKGROUNDS", (0, data_start), (-1, -1),
                      [colors.white, colors.HexColor("#F4F6FB")]))
    for j, col in enumerate(cols):
        if col.align == "right":
            style.append(("ALIGN", (j, data_start), (j, -1), "RIGHT"))
    table.setStyle(TableStyle(style))
    story.append(table)
    doc.build(story)
    return buf.getvalue()


def filename(result, ext):
    b = branding()
    stub = b["product"].replace(" ", "")
    stamp = datetime.utcnow().strftime("%Y%m%d")
    return f"{stub}_{result.query_id}_{stamp}.{ext}"
