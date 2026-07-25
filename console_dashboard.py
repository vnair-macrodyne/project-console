"""
console_dashboard.py
=====================
Project Console — Executive Dashboard BUILDER (Phase 1 delivery / last mile).

Turns the Layer-1 engine output into the Project Console **Executive Dashboard** — one
ranked row per project — and writes it as a styled sheet into a copy of the
management workbook. This is the "deliver the automated numbers where management
consumes them" step called for in CARPEDIA_PHASE_PLAN.md (Phase 1 gap #5).

WHERE EACH CELL COMES FROM (CARPEDIA_DASHBOARD_SPEC.md §2)
---------------------------------------------------------
  Block          Metric                         Source
  ────────────   ────────────────────────────   ───────────────────────────────
  Schedule       PO / Cust-Agreed / Planned      overlay  (manual — not in ETO)
                 ship dates, % Done
                 Slippage                         DERIVED  (Planned − Cust-Agreed)
  Budget         % Consumed (Labour, hours)       ETO      (Layer 1 LabPctHrs)   ← primary
                 % Consumed (Labour, $)           ETO      (Layer 1 LabPctCost)
                 % Consumed (Material)            ETO      (Layer 1 MatPct)
                 Run-out (Labour)                 DERIVED  (LabPctHrs ÷ %Done) or overlay
                 Run-out (Material)               overlay
  2-Week Delta   % Done delta                     overlay  (manual)
                 Labour Hrs (last 2 wks)          ETO      (Labor Data feed)
                 Material Spend delta             overlay  (manual today; Phase 3A)
  Labour         % consumed per discipline        ETO      (Layer 1 discipline df)
                 (PM/Mech/Hyd/Elec/Mfg)
  Procurement    LLTP line-item lateness counts   overlay  (manual today; Phase 3A)

KEY DISCOVERY BAKED IN: the Budget block and the per-discipline Labour block are
computed straight from ETO — Layer 1 pulls the *estimate* (denominator) from
ETO's estimating views, not the manual Budgets tab — so the only genuinely manual
input left is the schedule/procurement overlay. Lead with **hours %** as the
cross-project comparator (estimate-$ basis is inconsistent — see phase plan gap #3).

METHOD DISCIPLINE: writes to a NEW sheet `Executive Dashboard (Auto)` and leaves
the hand-built `Executive Dashboard` untouched, so the two can sit side by side
for one validation cycle before the manual one is retired (deprecate-then-delete).

Runnable end-to-end on the domain-joined box:
    python console_dashboard.py --workbook "Macrodyne Executive Dashboard.xlsx" \
                                 --overlay overlay.csv --projects 230219,240033,...
"""
from __future__ import annotations

import argparse
import datetime as _dt

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from console_feed import DISCIPLINES
from console_config import TENANT

# ─────────────────────────────────────────────────────────────────────────────
# Palette — header colour is tenant-branded (TENANT.header_color).
# ─────────────────────────────────────────────────────────────────────────────
C_HEADER   = TENANT.header_color   # metric header row (tenant brand)
C_GROUP    = "2E75B6"   # medium blue — block-group header row
C_TITLE    = TENANT.header_color
C_ALT      = "F2F4F8"   # zebra shading
C_GOOD     = "C6EFCE"   # ≤90% consumed / on-schedule green
C_WARN     = "FFEB9C"   # 90–100% amber
C_BAD      = "FFC7CE"   # >100% consumed / slipped red
WHITE      = "FFFFFF"
GREY       = "808080"

COMPANY = TENANT.company_name       # tenant profile, not a literal
AUTO_SHEET = "Executive Dashboard (Auto)"

# Discipline short labels for the Labour block header (5 real disciplines; Other
# is rolled into the summary but not shown as its own budget column).
_DISC_LABELS = {
    "Project Management":     "PM",
    "Mechanical Engineering": "Mechanical",
    "Hydraulic Engineering":  "Hydraulic",
    "Electrical Engineering": "Electrical",
    "Manufacturing":          "Manufacturing",
}
_DISC_ORDER = ["Project Management", "Mechanical Engineering", "Hydraulic Engineering",
               "Electrical Engineering", "Manufacturing"]

# ─────────────────────────────────────────────────────────────────────────────
# COLUMN CONTRACT
# Each entry: (key, header, block, kind)  where kind ∈ {text,int,pct,money,days,date}
# `block` drives the merged group-header row. Order here == column order.
# ─────────────────────────────────────────────────────────────────────────────
COLUMNS = [
    ("Rank",            "Rank",                  "",             "int"),
    ("ProjectID",       "Proj ID",               "",             "text"),
    ("Project",         "Project",               "",             "text"),
    # Schedule
    ("POShipDate",      "P.O. Ship",             "Schedule",     "date"),
    ("CustAgreedDate",  "Cust. Agreed",          "Schedule",     "date"),
    ("PlannedShipDate", "Planned Ship",          "Schedule",     "date"),
    ("SlippageDays",    "Slippage (d)",          "Schedule",     "days"),
    ("PctDone",         "% Done",                "Schedule",     "pct"),
    # Budget
    ("LabPctHrs",       "Labour % (hrs)",        "Budget",       "pct"),
    ("LabPctCost",      "Labour % ($)",          "Budget",       "pct"),
    ("RunoutLabour",    "Run-out Lab.",          "Budget",       "pct"),
    ("MatPct",          "Material %",            "Budget",       "pct"),
    ("RunoutMaterial",  "Run-out Mat.",          "Budget",       "pct"),
    # 2-Week Delta
    ("PctDoneDelta",    "Δ % Done",              "2-Week Delta", "pct"),
    ("LabHrs2wk",       "Δ Labour Hrs",          "2-Week Delta", "int"),
    ("MatSpend2wk",     "Δ Material $",          "2-Week Delta", "money"),
    # Labour by discipline (filled dynamically below)
    # Procurement
    ("TotalLineItems",  "Line Items",            "Procurement",  "int"),
    ("LLTPOrdered",     "LLTP Ord.",             "Procurement",  "int"),
    ("LLTPRelLate",     "LLTP Rel. Late",        "Procurement",  "int"),
    ("LLTPOrdLate",     "LLTP Ord. Late",        "Procurement",  "int"),
    ("LLTPDelLate",     "LLTP Del. Late",        "Procurement",  "int"),
    ("PartsRelLate",    "Parts Rel. Late",       "Procurement",  "int"),
    ("PartsOrdLate",    "Parts Ord. Late",       "Procurement",  "int"),
]


def _column_contract():
    """Insert the per-discipline Labour columns after the 2-Week Delta block."""
    cols = []
    inserted = False
    for entry in COLUMNS:
        if entry[2] == "Procurement" and not inserted:
            for disc in _DISC_ORDER:
                cols.append((f"disc::{disc}", _DISC_LABELS[disc], "Labour", "pct"))
            inserted = True
        cols.append(entry)
    return cols


# ─────────────────────────────────────────────────────────────────────────────
# Small numeric helpers
# ─────────────────────────────────────────────────────────────────────────────
def _num(x):
    try:
        if x is None or (isinstance(x, float) and pd.isna(x)):
            return None
        return float(x)
    except (TypeError, ValueError):
        return None


def _frac(x):
    """Normalise a %-done that may arrive as 0–1 or 0–100 into a 0–1 fraction."""
    v = _num(x)
    if v is None:
        return None
    return v / 100.0 if v > 1.0 else v


def _as_date(x):
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return None
    if isinstance(x, _dt.datetime):
        x = x.date()
    if isinstance(x, _dt.date):
        return None if x.year >= 2099 else x   # 2099 = PM placeholder for "TBD"
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
        try:
            d = _dt.datetime.strptime(str(x).strip(), fmt).date()
            return None if d.year >= 2099 else d
        except ValueError:
            continue
    return None


# ─────────────────────────────────────────────────────────────────────────────
# 2-week labour from the Labor Data feed (ETO)
# ─────────────────────────────────────────────────────────────────────────────
def two_week_labour_hours(labor_data: pd.DataFrame, as_of: _dt.date | None = None) -> dict:
    """
    Σ Actual Hours per ProjectID over the trailing 14 days (matches the dashboard's
    '2-Week Delta / Labor Hrs' cell). `labor_data` is the DataFrame from
    console_feed.query_labor_data(). Returns {ProjectID(str): hours}.
    """
    if labor_data is None or labor_data.empty:
        return {}
    df = labor_data.copy()
    df["Date"] = df["Date"].map(_as_date)
    if as_of is None:
        as_of = max(d for d in df["Date"] if d is not None)
    lo = as_of - _dt.timedelta(days=13)
    win = df[(df["Date"].map(lambda d: d is not None and lo <= d <= as_of))]
    grp = win.groupby(win["Project ID"].astype(str))["Actual Hours"].sum()
    return {k: round(float(v), 1) for k, v in grp.items()}


# ─────────────────────────────────────────────────────────────────────────────
# Assembly: Layer 1 output + overlay → Executive Dashboard rows
# ─────────────────────────────────────────────────────────────────────────────
def build_executive_dashboard(summary: pd.DataFrame,
                              disc: pd.DataFrame,
                              overlay: pd.DataFrame | None = None,
                              two_week: dict | None = None) -> pd.DataFrame:
    """
    Merge Layer-1 output with the manual overlay into one ranked row per project.

    summary  : console_engine.build_project_summary() output
    disc     : console_engine.build_discipline_labour() output (long form)
    overlay  : optional manual overlay, one row per ProjectID. Recognised columns:
               Rank, POShipDate, CustAgreedDate, PlannedShipDate, PctDone,
               RunoutLabour, RunoutMaterial, PctDoneDelta, MatSpend2wk,
               TotalLineItems, LLTPOrdered, LLTPRelLate, LLTPOrdLate, LLTPDelLate,
               PartsRelLate, PartsOrdLate. All optional; absent → blank.
    two_week : {ProjectID(str): 2-week labour hours} from two_week_labour_hours().

    Returns a DataFrame whose columns == [c[0] for c in _column_contract()],
    sorted by Rank (projects without a rank sort last, by descending Labour %).
    """
    ov = _index_overlay(overlay)
    tw = two_week or {}

    # pivot ACTUAL discipline hours (ETO) to wide: ProjectID → {discipline: act_hours}
    disc_act = {}
    if disc is not None and not disc.empty:
        hrs_col = "ActHours" if "ActHours" in disc.columns else "PctConsumed"
        for pid, g in disc.groupby(disc["ProjectID"].astype(str)):
            disc_act[pid] = dict(zip(g["Discipline"], g[hrs_col]))

    rows = []
    for _, s in summary.iterrows():
        pid = str(int(s["ProjectID"]))
        o = ov.get(pid, {})
        planned = _as_date(o.get("PlannedShipDate"))
        agreed = _as_date(o.get("CustAgreedDate"))
        pct_done = _frac(o.get("PctDone"))
        lab_pct_hrs = _num(s.get("LabPctHrs"))

        row = {
            "Rank": _num(o.get("Rank")),
            "ProjectID": pid,
            "Project": s.get("Project") if pd.notna(s.get("Project")) else pid,
            # Schedule (manual overlay + one derived)
            "POShipDate":      _as_date(o.get("POShipDate")),
            "CustAgreedDate":  agreed,
            "PlannedShipDate": planned,
            "SlippageDays":    (planned - agreed).days if (planned and agreed) else None,
            "PctDone":         pct_done,
            # Budget (ETO — Layer 1)
            "LabPctHrs":  lab_pct_hrs,
            "LabPctCost": _num(s.get("LabPctCost")),
            "MatPct":     _num(s.get("MatPct")),
            # Run-out labour: prefer the PM's own run-out (PM Entries), else derive.
            "RunoutLabour":   _num(o.get("RunoutLabour")) if o.get("RunoutLabour") is not None
                              else ((lab_pct_hrs / pct_done) if (lab_pct_hrs is not None and pct_done) else None),
            "RunoutMaterial": _num(o.get("RunoutMaterial")),
            # 2-week
            "PctDoneDelta": _frac(o.get("PctDoneDelta")),
            "LabHrs2wk":    tw.get(pid),
            "MatSpend2wk":  _num(o.get("MatSpend2wk")),
            # Procurement (manual overlay)
            "TotalLineItems": _num(o.get("TotalLineItems")),
            "LLTPOrdered":    _num(o.get("LLTPOrdered")),
            "LLTPRelLate":    _num(o.get("LLTPRelLate")),
            "LLTPOrdLate":    _num(o.get("LLTPOrdLate")),
            "LLTPDelLate":    _num(o.get("LLTPDelLate")),
            "PartsRelLate":   _num(o.get("PartsRelLate")),
            "PartsOrdLate":   _num(o.get("PartsOrdLate")),
        }
        # Labour by discipline on HOURS: actual hrs (ETO) ÷ budget hrs (Budgets tab).
        # Blank when the PM hasn't entered that discipline's budget yet.
        da = disc_act.get(pid, {})
        for d in _DISC_ORDER:
            act = _num(da.get(d))
            bud = _num(o.get(f"BudgetHrs::{d}"))
            row[f"disc::{d}"] = (act / bud) if (act is not None and bud) else None
        rows.append(row)

    df = pd.DataFrame(rows, columns=[c[0] for c in _column_contract()])
    # rank sort: ranked rows first (asc), unranked after by descending labour %
    df["_hasrank"] = df["Rank"].notna()
    df = df.sort_values(
        by=["_hasrank", "Rank", "LabPctHrs"],
        ascending=[False, True, False],
        na_position="last",
    ).drop(columns="_hasrank").reset_index(drop=True)
    # auto-number blank ranks so every row shows a rank
    df["Rank"] = range(1, len(df) + 1)
    return df


def _index_overlay(overlay):
    """Normalise the overlay DataFrame to {ProjectID(str): {col: val}}."""
    if overlay is None or len(overlay) == 0:
        return {}
    df = overlay.copy()
    # accept either 'ProjectID' or 'Project ID'
    key = "ProjectID" if "ProjectID" in df.columns else (
        "Project ID" if "Project ID" in df.columns else None)
    if key is None:
        raise ValueError("overlay must have a 'ProjectID' (or 'Project ID') column")
    out = {}
    for _, r in df.iterrows():
        pid = str(r[key]).split(".")[0].strip()
        out[pid] = {k: r[k] for k in df.columns if k != key}
    return out


# ─────────────────────────────────────────────────────────────────────────────
# Writer: styled 'Executive Dashboard (Auto)' sheet into a copy of the workbook
# ─────────────────────────────────────────────────────────────────────────────
def _fmt_code(kind):
    return {
        "int":   "#,##0",
        "pct":   "0.0%",
        "money": "$#,##0",
        "days":  "#,##0",
        "date":  "yyyy-mm-dd",
        "text":  "@",
    }[kind]


def _consumption_fill(kind, val):
    """Traffic-light fill for % consumed and slippage cells (subtle, scannable)."""
    if val is None:
        return None
    if kind == "pct":
        if val > 1.0:
            return C_BAD
        if val >= 0.9:
            return C_WARN
        return None
    if kind == "days":
        if val > 0:
            return C_BAD
        return None
    return None


def write_executive_dashboard(df: pd.DataFrame,
                             output_path: str,
                             template_path: str | None = None,
                             as_of: _dt.date | None = None,
                             traffic_light: bool = True):
    """
    Write the assembled dashboard DataFrame to `output_path`.

    If `template_path` is given, the existing workbook is loaded and every other
    tab is preserved (the manual `Executive Dashboard` stays intact); the auto
    sheet is (re)created as `Executive Dashboard (Auto)`. If no template, a fresh
    workbook is created (used by the synthetic test harness).
    """
    contract = _column_contract()
    ncol = len(contract)

    if template_path:
        wb = load_workbook(template_path)
        if AUTO_SHEET in wb.sheetnames:
            del wb[AUTO_SHEET]
        ws = wb.create_sheet(AUTO_SHEET)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = AUTO_SHEET

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    as_of = as_of or _dt.date.today()

    # ── Title block (rows 1–3) ─────────────────────────────────────────────
    ws.cell(1, 1, COMPANY).font = Font(name="Helvetica", size=12, bold=True)
    ws.cell(2, 1, f"{TENANT.product_name} — Executive Dashboard — automated from ETO").font = \
        Font(name="Helvetica", size=10, bold=True, color=C_TITLE)
    ws.cell(3, 1, f"As of {as_of:%Y-%m-%d}  ·  Budget & Labour blocks: ETO (Layer 1)  "
                  f"·  Schedule & Procurement: manual overlay  ·  lead metric = Labour % (hrs)"
            ).font = Font(name="Helvetica", size=8, italic=True, color=GREY)

    GROUP_ROW = 5
    HEAD_ROW = 6
    DATA_ROW = 7

    # ── Group header (merged per block) ────────────────────────────────────
    group_fill = PatternFill("solid", fgColor=C_GROUP)
    group_font = Font(name="Helvetica", size=9, bold=True, color=WHITE)
    col = 1
    while col <= ncol:
        block = contract[col - 1][2]
        span = 1
        while col + span <= ncol and contract[col + span - 1][2] == block:
            span += 1
        c = ws.cell(GROUP_ROW, col, block if block else "")
        c.fill = group_fill
        c.font = group_font
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
        if span > 1:
            ws.merge_cells(start_row=GROUP_ROW, start_column=col,
                           end_row=GROUP_ROW, end_column=col + span - 1)
            for k in range(1, span):
                cc = ws.cell(GROUP_ROW, col + k)
                cc.fill = group_fill
                cc.border = border
        col += span

    # ── Metric header ──────────────────────────────────────────────────────
    head_fill = PatternFill("solid", fgColor=C_HEADER)
    head_font = Font(name="Helvetica", size=8, bold=True, color=WHITE)
    for i, (_key, header, _block, _kind) in enumerate(contract, 1):
        c = ws.cell(HEAD_ROW, i, header)
        c.fill = head_fill
        c.font = head_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    # ── Data rows ──────────────────────────────────────────────────────────
    for r, (_, rowdata) in enumerate(df.iterrows(), DATA_ROW):
        zebra = PatternFill("solid", fgColor=C_ALT) if (r - DATA_ROW) % 2 else None
        for i, (key, _header, _block, kind) in enumerate(contract, 1):
            val = rowdata[key]
            if isinstance(val, float) and pd.isna(val):
                val = None
            c = ws.cell(r, i, val)
            c.border = border
            c.font = Font(name="Helvetica", size=8)
            if kind == "text":
                c.alignment = Alignment(horizontal="left", vertical="center")
            else:
                c.alignment = Alignment(horizontal="center", vertical="center")
            if val is not None:
                c.number_format = _fmt_code(kind)
            # traffic-light beats zebra where it applies
            fill = _consumption_fill(kind, val) if traffic_light else None
            if fill:
                c.fill = PatternFill("solid", fgColor=fill)
            elif zebra:
                c.fill = zebra

    # ── Layout niceties ────────────────────────────────────────────────────
    ws.freeze_panes = ws.cell(DATA_ROW, 4)  # freeze title/headers + Rank+ProjID+Project cols
    widths = {"text": 30, "date": 12, "pct": 11, "money": 12, "int": 9, "days": 10}
    for i, (_k, _h, _b, kind) in enumerate(contract, 1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(kind, 11)

    last = DATA_ROW + len(df) - 1
    foot = last + 2
    ws.cell(foot, 1, f"Generated {_dt.datetime.now():%Y-%m-%d %H:%M} · {TENANT.confidential_footer} "
                     f"· Cost basis: applied rate (HourTime×HourRate×HourFactor)").font = \
        Font(name="Helvetica", size=7, italic=True, color=GREY)

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    wb.save(output_path)
    return output_path


# ─────────────────────────────────────────────────────────────────────────────
# Overlay CSV helpers
# ─────────────────────────────────────────────────────────────────────────────
OVERLAY_TEMPLATE_COLUMNS = [
    "ProjectID", "Rank", "POShipDate", "CustAgreedDate", "PlannedShipDate",
    "PctDone", "RunoutLabour", "RunoutMaterial", "PctDoneDelta", "MatSpend2wk",
    "TotalLineItems", "LLTPOrdered", "LLTPRelLate", "LLTPOrdLate", "LLTPDelLate",
    "PartsRelLate", "PartsOrdLate",
]


def load_overlay_csv(path):
    return pd.read_csv(path, dtype={"ProjectID": str})


def write_overlay_template(path, project_ids):
    rows = [{**{c: "" for c in OVERLAY_TEMPLATE_COLUMNS}, "ProjectID": str(int(p))}
            for p in project_ids]
    pd.DataFrame(rows, columns=OVERLAY_TEMPLATE_COLUMNS).to_csv(path, index=False)
    return path


# ─────────────────────────────────────────────────────────────────────────────
# End-to-end runner (domain-joined machine, live ETO)
# ─────────────────────────────────────────────────────────────────────────────
def _merge_overlays(frames):
    """Combine multiple overlay frames into one row-per-ProjectID (first non-null wins)."""
    frames = [f for f in frames if f is not None and len(f)]
    if not frames:
        return None
    df = pd.concat(frames, ignore_index=True)
    key = "ProjectID" if "ProjectID" in df.columns else "Project ID"
    df[key] = df[key].astype(str).str.split(".").str[0].str.strip()
    return df.groupby(key, as_index=False).first()


def run(project_ids, workbook_path=None, overlay_path=None, output_path=None,
        two_week_days=14, standalone=False, make_budgets=False):
    """
    Live build: connect → Layer 1 → 2-week labour → assemble → write.

    Manual inputs (per-discipline budget hours + ship dates + % done) are read from
    the pack's `Budgets (Input)` sheet — the PM entry point — with an optional CSV
    overlay layered on for the remaining PM-Entries fields (procurement, 2-wk deltas).

    make_budgets=True instead (re)writes the SEEDED Budgets (Input) sheet from ETO
    and exits — the PM then fills it before the next dashboard run.

    Embeds into a COPY of `workbook_path` (preserving other tabs); `standalone=True`
    or a missing workbook falls back to a fresh one-sheet file.
    """
    import os
    from console_engine import project_scorecard, query_two_week_labour_hours, _connect
    from console_budgets import (write_budgets_input, read_budgets_input,
                                  seed_columns_present, BUDGETS_SHEET)

    today = _dt.date.today()
    conn = _connect()
    try:
        cur = conn.cursor()
        summary, disc = project_scorecard(cur, project_ids)
        tw = query_two_week_labour_hours(cur, project_ids, as_of=today, days=two_week_days)
    finally:
        conn.close()

    has_wb = bool(workbook_path) and os.path.exists(workbook_path)

    # ── make-budgets: emit/refresh the seeded PM entry form and stop ────────────
    if make_budgets:
        if not seed_columns_present(summary):
            raise RuntimeError("summary is missing ETO estimate-hours columns for seeding")
        out = output_path or (workbook_path.replace(".xlsx", "_BUDGETS.xlsx")
                              if has_wb else "Project Console_Budgets_Input.xlsx")
        write_budgets_input(out, summary, template_path=workbook_path if has_wb else None)
        print(f"Wrote seeded '{BUDGETS_SHEET}' for {len(summary)} projects → {out}\n"
              f"Fill the yellow cells (split Engineering into Mech/Hyd/Elec + ship dates), "
              f"then run the dashboard.")
        return out

    # ── overlay: read the manual entry points ───────────────────────────────────
    # A REAL management pack (has 'Budgets' + 'PM Entries') is read via console_pack
    # and NEVER written into (charts/tables don't survive openpyxl round-trip) — the
    # dashboard goes to a separate file. Otherwise fall back to my seeded Budgets (Input).
    overlays = []
    is_real_pack = False
    if has_wb:
        from openpyxl import load_workbook as _lw
        names = _lw(workbook_path, read_only=True).sheetnames
        if "Budgets" in names and "PM Entries" in names:
            is_real_pack = True
            from console_pack import read_pack_overlay, selected_week
            po = read_pack_overlay(workbook_path)
            if po is not None:
                overlays.append(po)
            print(f"Read the real pack: Budgets + PM Entries (week {selected_week(workbook_path)}).")
        else:
            b = read_budgets_input(workbook_path)
            if b is not None:
                overlays.append(b)
            else:
                print(f"NOTE: no Budgets source in the workbook — discipline % and schedule "
                      f"will be blank.")
    if overlay_path:
        overlays.append(load_overlay_csv(overlay_path))
    overlay = _merge_overlays(overlays)

    df = build_executive_dashboard(summary, disc, overlay, two_week=tw)

    # A real pack is never overwritten; always emit a separate dashboard file.
    template = None if (is_real_pack or standalone) else (workbook_path if has_wb else None)
    if not template and workbook_path and not standalone and not has_wb:
        print(f"WARNING: workbook not found:\n    {workbook_path}\n"
              f"Falling back to a STANDALONE one-sheet workbook.")
    if output_path is None:
        output_path = (workbook_path.replace(".xlsx", "_AUTO.xlsx")
                       if template else "Project Console_Executive_Dashboard_AUTO.xlsx")

    write_executive_dashboard(df, output_path, template_path=template, as_of=today)
    if is_real_pack:
        where = "separate file (pack left untouched — preserves charts/tables)"
    elif template:
        where = "added to a copy of the pack"
    else:
        where = "standalone workbook"
    print(f"Wrote '{AUTO_SHEET}' for {len(df)} projects ({where}) → {output_path}")
    return output_path


if __name__ == "__main__":
    ap = argparse.ArgumentParser(description="Build the Project Console — Executive Dashboard from ETO.")
    ap.add_argument("--workbook", help="Path to the management workbook (.xlsx) to embed the "
                                       "sheet into. Omit (or use --standalone) for a fresh one-sheet file.")
    ap.add_argument("--projects", required=True, help="Comma-separated ProjectIDs")
    ap.add_argument("--overlay", help="Manual overlay CSV (schedule/procurement)")
    ap.add_argument("--output", help="Output workbook path (default: <workbook>_AUTO.xlsx, "
                                     "or Project Console_Executive_Dashboard_AUTO.xlsx standalone)")
    ap.add_argument("--standalone", action="store_true",
                    help="Generate a fresh one-sheet workbook; do not load/require a source workbook")
    ap.add_argument("--make-budgets", action="store_true",
                    help="(Re)write the seeded 'Budgets (Input)' PM entry sheet from ETO and exit")
    ap.add_argument("--make-overlay-template", action="store_true",
                    help="Just write an overlay CSV template for --projects and exit")
    args = ap.parse_args()

    pids = [int(p) for p in args.projects.split(",") if p.strip()]
    if args.make_overlay_template:
        out = args.overlay or "overlay_template.csv"
        write_overlay_template(out, pids)
        print(f"Wrote overlay template → {out}")
    else:
        run(pids, args.workbook, args.overlay, args.output,
            standalone=args.standalone, make_budgets=args.make_budgets)
